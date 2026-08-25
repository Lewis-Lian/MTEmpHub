from __future__ import annotations

from datetime import datetime

import openpyxl
from flask import jsonify, request

from routes.auth_helpers import admin_required
from models.daily_attendance_override import DailyAttendanceOverride


def _requested_emp_ids() -> list[int]:
    raw_values = request.args.getlist("emp_ids")
    if not raw_values:
        raw = (request.args.get("emp_ids") or "").strip()
        raw_values = raw.split(",") if raw else []
    emp_ids: list[int] = []
    seen: set[int] = set()
    for raw in raw_values:
        for part in str(raw or "").split(","):
            text = part.strip()
            if not text or not text.isdigit():
                continue
            emp_id = int(text)
            if emp_id in seen:
                continue
            seen.add(emp_id)
            emp_ids.append(emp_id)
    return emp_ids


def manager_attendance_override_record_api():
    from routes import admin_core as admin_module

    emp_id = request.args.get("emp_id", type=int) or 0
    month = admin_module._validate_month(request.args.get("month"))
    if not emp_id or not month:
        return jsonify({"error": "请选择管理人员和有效月份"}), 400
    payload, status = admin_module._manager_attendance_response(emp_id, month)
    return jsonify(payload), status


def manager_attendance_override_list_api():
    from routes import admin_core as admin_module

    month = admin_module._validate_month(request.args.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    payload, status = admin_module._manager_attendance_list_response(_requested_emp_ids(), month)
    return jsonify(payload), status


def save_manager_attendance_override_record_api():
    from routes import admin_core as admin_module

    data = request.json or {}
    emp_id = int(data.get("emp_id") or 0)
    month = admin_module._validate_month(data.get("month"))
    if not emp_id or not month:
        return jsonify({"error": "请选择管理人员和有效月份"}), 400
    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "保存管理人员考勤修正")
    if locked_error:
        return locked_error
    employee = admin_module.db.session.get(admin_module.Employee, emp_id)
    if not employee or not employee.is_manager:
        return jsonify({"error": "employee is not manager"}), 400

    values: dict[str, float | int | None] = {}
    for key in (
        "attendance_days",
        "injury_days",
        "business_trip_days",
        "marriage_days",
        "funeral_days",
    ):
        value, error = admin_module._nullable_float(data, key)
        if error:
            return jsonify({"error": error}), 400
        values[key] = value
    late_value, error = admin_module._nullable_int(data, "late_early_minutes")
    if error:
        return jsonify({"error": error}), 400
    values["late_early_minutes"] = late_value

    row = admin_module.ManagerAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    before_values = admin_module._override_state_from_row(
        row, admin_module._MANAGER_ATTENDANCE_OVERRIDE_FIELDS
    )
    after_values = dict(values)
    after_values["remark"] = (data.get("remark") or "").strip()
    if admin_module._has_override_state_changes(before_values, after_values):
        if not row:
            row = admin_module.ManagerAttendanceOverride(emp_id=emp_id, month=month)
            admin_module.db.session.add(row)
        for key, value in values.items():
            setattr(row, key, value)
        row.remark = after_values["remark"]
        row.updated_by = admin_module.g.current_user.id
        admin_module._record_override_history(
            "manager", emp_id, month, "manual_save", before_values, after_values
        )
        admin_module.db.session.commit()
        # 考勤天数修正会影响加班天数，需同步加班查询页数据源（manager_month_stats）
        admin_module._sync_manager_month_stats(month)

    payload, status = admin_module._manager_attendance_response(emp_id, month)
    return jsonify(payload), status


def delete_manager_attendance_override_record_api():
    from routes import admin_core as admin_module

    emp_id = request.args.get("emp_id", type=int) or 0
    month = admin_module._validate_month(request.args.get("month"))
    if not emp_id or not month:
        return jsonify({"error": "请选择管理人员和有效月份"}), 400
    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "清空管理人员考勤修正")
    if locked_error:
        return locked_error
    row = admin_module.ManagerAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    if row:
        before_values = admin_module._override_state_from_row(
            row, admin_module._MANAGER_ATTENDANCE_OVERRIDE_FIELDS
        )
        after_values = admin_module._override_state_from_row(
            None, admin_module._MANAGER_ATTENDANCE_OVERRIDE_FIELDS
        )
        admin_module._record_override_history("manager", emp_id, month, "clear", before_values, after_values)
        admin_module.db.session.delete(row)
        admin_module.db.session.commit()
        # 删除修正后出勤天数恢复，需同步重算加班查询页数据源
        admin_module._sync_manager_month_stats(month)
    payload, status = admin_module._manager_attendance_response(emp_id, month)
    return jsonify(payload), status


def manager_attendance_override_history_api():
    from routes import admin_core as admin_module

    month = admin_module._validate_month(request.args.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    return jsonify({"rows": admin_module._history_rows_for_month("manager", month)})


def employee_attendance_override_record_api():
    from routes import admin_core as admin_module

    emp_id = request.args.get("emp_id", type=int) or 0
    month = admin_module._validate_month(request.args.get("month"))
    if not emp_id or not month:
        return jsonify({"error": "请选择员工和有效月份"}), 400
    payload, status = admin_module._employee_override_response(emp_id, month)
    return jsonify(payload), status


def employee_attendance_override_list_api():
    from routes import admin_core as admin_module

    month = admin_module._validate_month(request.args.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    payload, status = admin_module._employee_override_list_response(_requested_emp_ids(), month)
    return jsonify(payload), status


def save_employee_attendance_override_record_api():
    from routes import admin_core as admin_module

    data = request.json or {}
    emp_id = int(data.get("emp_id") or 0)
    month = admin_module._validate_month(data.get("month"))
    if not emp_id or not month:
        return jsonify({"error": "请选择员工和有效月份"}), 400
    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "保存员工考勤修正")
    if locked_error:
        return locked_error
    employee = admin_module.db.session.get(admin_module.Employee, emp_id)
    if not employee or employee.is_manager:
        return jsonify({"error": "员工不存在或是管理人员"}), 400

    values: dict[str, float | int | None] = {}
    for key in ("attendance_days", "actual_attendance_days", "work_hours"):
        value, error = admin_module._nullable_float(data, key)
        if error:
            return jsonify({"error": error}), 400
        values[key] = value
    for key in ("half_days", "late_early_minutes"):
        value, error = admin_module._nullable_int(data, key)
        if error:
            return jsonify({"error": error}), 400
        values[key] = value

    row = admin_module.EmployeeAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    updates = dict(values)
    updates["remark"] = (data.get("remark") or "").strip()
    if admin_module._apply_employee_override_updates(row, emp_id, month, updates, "manual_save"):
        admin_module.db.session.commit()

    payload, status = admin_module._employee_override_response(emp_id, month)
    return jsonify(payload), status


def delete_employee_attendance_override_record_api():
    from routes import admin_core as admin_module

    emp_id = request.args.get("emp_id", type=int) or 0
    month = admin_module._validate_month(request.args.get("month"))
    if not emp_id or not month:
        return jsonify({"error": "请选择员工和有效月份"}), 400
    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "清空员工考勤修正")
    if locked_error:
        return locked_error
    row = admin_module.EmployeeAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    if row:
        before_values = admin_module._override_state_from_row(row, admin_module._EMPLOYEE_OVERRIDE_FIELDS)
        after_values = admin_module._override_state_from_row(None, admin_module._EMPLOYEE_OVERRIDE_FIELDS)
        admin_module._record_override_history("employee", emp_id, month, "clear", before_values, after_values)
        admin_module.db.session.delete(row)
        admin_module.db.session.commit()
    payload, status = admin_module._employee_override_response(emp_id, month)
    return jsonify(payload), status


def employee_attendance_override_history_api():
    from routes import admin_core as admin_module

    month = admin_module._validate_month(request.args.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    return jsonify({"rows": admin_module._history_rows_for_month("employee", month)})


# ---------------------------------------------------------------- 逐日考勤修正


def _daily_record_response(employee, month: str):
    """逐日修正保存/清除后的统一响应：{calendar, row}——前端据此同时刷新日历弹窗与外层列表行。"""
    from routes import admin_core as admin_module
    from routes.query_core import _build_attendance_calendar_payload

    if employee.is_manager:
        row_payload, _ = admin_module._manager_attendance_response(employee.id, month)
    else:
        row_payload, _ = admin_module._employee_override_response(employee.id, month)
    return jsonify(
        {
            "calendar": _build_attendance_calendar_payload(employee, month),
            "row": row_payload,
        }
    )


def _parse_daily_override_payload(data: dict, employee) -> tuple[dict, tuple | None]:
    """校验并解析逐日修正字段。返回 (values, error_response)。"""
    from routes import admin_core as admin_module
    from services.daily_override_service import daily_statuses_for

    values: dict[str, object] = {}

    status = str(data.get("status") or "").strip()
    if status and status not in daily_statuses_for(employee):
        return values, (jsonify({"error": f"无效的考勤状态：{status}"}), 400)
    values["status"] = status or None

    for key, parser in (("work_hours", admin_module._nullable_float), ("late_minutes", admin_module._nullable_int), ("early_leave_minutes", admin_module._nullable_int)):
        parsed, error = parser(data, key)
        if error:
            return values, (jsonify({"error": error}), 400)
        values[key] = parsed

    flag = data.get("is_evening_overtime")
    if flag in (None, ""):
        values["is_evening_overtime"] = None
    elif isinstance(flag, bool):
        values["is_evening_overtime"] = flag
    else:
        values["is_evening_overtime"] = str(flag).strip().lower() in ("true", "1", "yes", "是")

    values["remark"] = (data.get("remark") or "").strip()
    return values, None


def daily_attendance_override_calendar_api():
    from routes import admin_core as admin_module
    from routes.query_core import _build_attendance_calendar_payload

    emp_id = request.args.get("emp_id", type=int) or 0
    month = admin_module._validate_month(request.args.get("month"))
    if not emp_id or not month:
        return jsonify({"error": "请选择员工和有效月份"}), 400
    employee = admin_module.db.session.get(admin_module.Employee, emp_id)
    if not employee:
        return jsonify({"error": "员工不存在"}), 400
    return jsonify(_build_attendance_calendar_payload(employee, month))


def save_daily_attendance_override_record_api():
    from routes import admin_core as admin_module
    from services.daily_override_service import DAILY_OVERRIDE_FIELDS

    data = request.json or {}
    emp_id = int(data.get("emp_id") or 0)
    month = admin_module._validate_month(data.get("month"))
    date_text = str(data.get("date") or "").strip()
    if not emp_id or not month or not date_text:
        return jsonify({"error": "请选择员工、月份和日期"}), 400
    try:
        record_date = datetime.strptime(date_text, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "无效的日期"}), 400
    if record_date.strftime("%Y-%m") != month:
        return jsonify({"error": "日期与月份不一致"}), 400

    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "保存逐日考勤修正")
    if locked_error:
        return locked_error
    employee = admin_module.db.session.get(admin_module.Employee, emp_id)
    if not employee:
        return jsonify({"error": "员工不存在"}), 400

    values, error_response = _parse_daily_override_payload(data, employee)
    if error_response:
        return error_response

    row = DailyAttendanceOverride.query.filter_by(emp_id=emp_id, record_date=record_date).first()
    before_values = {"record_date": date_text}
    for field in DAILY_OVERRIDE_FIELDS:
        before_values[field] = getattr(row, field) if row else None
    if row:
        before_values["remark"] = row.remark or ""
    after_values = dict(before_values)
    after_values.update(values)
    if admin_module._has_override_state_changes(before_values, after_values):
        if not row:
            row = DailyAttendanceOverride(emp_id=emp_id, record_date=record_date)
            admin_module.db.session.add(row)
        for key, value in values.items():
            setattr(row, key, value)
        row.updated_by = admin_module.g.current_user.id
        admin_module._record_override_history("daily", emp_id, month, "manual_save", before_values, after_values)
        admin_module.db.session.commit()
        if employee.is_manager:
            # 逐日修正影响出勤/假种天数，需同步重算加班查询页数据源
            admin_module._sync_manager_month_stats(month)

    return _daily_record_response(employee, month)


def delete_daily_attendance_override_record_api():
    from routes import admin_core as admin_module
    from services.daily_override_service import DAILY_OVERRIDE_FIELDS

    emp_id = request.args.get("emp_id", type=int) or 0
    date_text = str(request.args.get("date") or "").strip()
    if not emp_id or not date_text:
        return jsonify({"error": "请选择员工和日期"}), 400
    try:
        record_date = datetime.strptime(date_text, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "无效的日期"}), 400
    month = record_date.strftime("%Y-%m")

    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "清除逐日考勤修正")
    if locked_error:
        return locked_error
    employee = admin_module.db.session.get(admin_module.Employee, emp_id)
    if not employee:
        return jsonify({"error": "员工不存在"}), 400

    row = DailyAttendanceOverride.query.filter_by(emp_id=emp_id, record_date=record_date).first()
    if row:
        before_values = {"record_date": date_text}
        for field in DAILY_OVERRIDE_FIELDS:
            before_values[field] = getattr(row, field)
        before_values["remark"] = row.remark or ""
        after_values = dict(before_values)
        for field in DAILY_OVERRIDE_FIELDS:
            after_values[field] = None
        after_values["remark"] = ""
        admin_module._record_override_history("daily", emp_id, month, "clear", before_values, after_values)
        admin_module.db.session.delete(row)
        admin_module.db.session.commit()
        if employee.is_manager:
            admin_module._sync_manager_month_stats(month)

    return _daily_record_response(employee, month)



# ---------------------------------------------------------------- 迟到冲抵


def late_offset_candidates_api():
    from routes import admin_core as admin_module
    from services.late_offset_service import late_offset_candidates

    month = admin_module._validate_month(request.args.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    emp_ids = _requested_emp_ids() if request.args.get("emp_ids") else None
    return jsonify({"month": month, "rows": late_offset_candidates(month, emp_ids)})


def late_offset_confirm_api():
    from routes import admin_core as admin_module
    from services.daily_override_service import DAILY_OVERRIDE_FIELDS
    from services.late_offset_service import confirm_late_offset, late_offset_candidates

    data = request.json or {}
    month = admin_module._validate_month(data.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    items = data.get("items")
    if not isinstance(items, list) or not items:
        return jsonify({"error": "请选择要冲抵的记录"}), 400

    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "迟到冲抵确认")
    if locked_error:
        return locked_error

    confirmed: list[dict] = []
    skipped: list[dict] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        emp_id = 0
        try:
            emp_id = int(item.get("emp_id") or 0)
        except (TypeError, ValueError):
            pass
        date_text = str(item.get("date") or "").strip()
        offset_minutes = None
        if item.get("offset_minutes") not in (None, ""):
            try:
                offset_minutes = int(item.get("offset_minutes"))
            except (TypeError, ValueError):
                skipped.append({"emp_id": emp_id, "date": date_text, "error": "冲抵分钟须为整数"})
                continue
        if not emp_id or not date_text:
            skipped.append({"emp_id": emp_id, "date": date_text, "error": "参数不完整"})
            continue
        try:
            record_date = datetime.strptime(date_text, "%Y-%m-%d").date()
        except ValueError:
            skipped.append({"emp_id": emp_id, "date": date_text, "error": "无效的日期"})
            continue
        if record_date.strftime("%Y-%m") != month:
            skipped.append({"emp_id": emp_id, "date": date_text, "error": "日期与月份不一致"})
            continue

        row = DailyAttendanceOverride.query.filter_by(emp_id=emp_id, record_date=record_date).first()
        before_values = {"record_date": date_text}
        for field in DAILY_OVERRIDE_FIELDS:
            before_values[field] = getattr(row, field) if row else None
        if row:
            before_values["remark"] = row.remark or ""
        try:
            result = confirm_late_offset(emp_id, record_date, offset_minutes=offset_minutes)
        except ValueError as exc:
            skipped.append({"emp_id": emp_id, "date": date_text, "error": str(exc)})
            continue
        updated_row = DailyAttendanceOverride.query.filter_by(
            emp_id=emp_id, record_date=record_date
        ).first()
        after_values = dict(before_values)
        for field in DAILY_OVERRIDE_FIELDS:
            after_values[field] = getattr(updated_row, field) if updated_row else None
        after_values["remark"] = (updated_row.remark or "") if updated_row else ""
        admin_module._record_override_history(
            "daily", emp_id, month, "late_offset", before_values, after_values
        )
        admin_module.db.session.commit()
        confirmed.append(result)

    if confirmed:
        admin_module._sync_manager_month_stats(month)
    return jsonify(
        {
            "month": month,
            "confirmed": confirmed,
            "skipped": skipped,
            "rows": late_offset_candidates(month),
        }
    )


@admin_required
def download_manager_attendance_override_template():
    from routes import admin_core as admin_module
    month = admin_module._validate_month(request.args.get("month")) or datetime.now().strftime("%Y-%m")
    return admin_module._override_workbook_response(
        admin_module._build_manager_override_export_workbook(month, include_real_rows=False),
        "管理人员考勤修正导入示例.xlsx",
    )


@admin_required
def export_manager_attendance_overrides():
    from routes import admin_core as admin_module
    month = admin_module._validate_month(request.args.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    return admin_module._override_workbook_response(
        admin_module._build_manager_override_export_workbook(month, include_real_rows=True),
        f"管理人员考勤修正导出_{month}.xlsx",
    )


@admin_required
def import_manager_attendance_overrides():
    from routes import admin_core as admin_module
    month = admin_module._validate_month(request.form.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "导入管理人员考勤修正")
    if locked_error:
        return locked_error
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "请选择导入文件"}), 400
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    if not rows:
        return jsonify({"error": "empty file"}), 400
    header_idx, header_map = admin_module._parse_header_row(rows, ["月份", "工号", "姓名", "出勤天数", "备注"])
    required = ["月份", "工号", "姓名", "出勤天数", "工伤", "出差", "婚假", "丧假", "迟到早退", "备注"]
    missing = [key for key in required if key not in header_map]
    if missing:
        return jsonify({"error": f"缺少列：{', '.join(missing)}"}), 400
    success_count = skipped_count = failed_count = changed_count = 0
    errors: list[str] = []
    for row_index, raw in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        row_month = (
            str(raw[header_map["月份"]]).strip()
            if header_map["月份"] < len(raw) and raw[header_map["月份"]] is not None
            else ""
        )
        emp_no = (
            str(raw[header_map["工号"]]).strip()
            if header_map["工号"] < len(raw) and raw[header_map["工号"]] is not None
            else ""
        )
        if not row_month and not emp_no:
            skipped_count += 1
            continue
        if row_month != month:
            failed_count += 1
            errors.append(f"第 {row_index} 行：月份 {row_month or '空'} 与当前月份 {month} 不一致")
            continue
        employee = admin_module.Employee.query.filter_by(emp_no=emp_no).first()
        if not employee or not employee.is_manager:
            failed_count += 1
            errors.append(f"第 {row_index} 行：工号 {emp_no or '空'} 未找到管理人员")
            continue
        updates: dict[str, object] = {}
        for key in (
            "attendance_days",
            "injury_days",
            "business_trip_days",
            "marriage_days",
            "funeral_days",
        ):
            label = admin_module._MANAGER_OVERRIDE_LABELS[key]
            value = raw[header_map[label]] if header_map[label] < len(raw) else None
            parsed, error = admin_module._nullable_float({key: value}, key)
            if error:
                failed_count += 1
                errors.append(f"第 {row_index} 行：{error}")
                updates = {}
                break
            if value not in (None, ""):
                updates[key] = parsed
        if not updates and failed_count and errors and errors[-1].startswith(f"第 {row_index} 行"):
            continue
        late_value = raw[header_map["迟到早退"]] if header_map["迟到早退"] < len(raw) else None
        parsed_late, error = admin_module._nullable_int(
            {"late_early_minutes": late_value}, "late_early_minutes"
        )
        if error:
            failed_count += 1
            errors.append(f"第 {row_index} 行：{error}")
            continue
        if late_value not in (None, ""):
            updates["late_early_minutes"] = parsed_late
        remark_value = raw[header_map["备注"]] if header_map["备注"] < len(raw) else None
        if remark_value not in (None, ""):
            updates["remark"] = str(remark_value).strip()
        row_obj = admin_module.ManagerAttendanceOverride.query.filter_by(
            emp_id=employee.id, month=month
        ).first()
        changed = admin_module._apply_manager_override_updates(
            row_obj, employee.id, month, updates, "import", file.filename
        )
        if changed:
            success_count += 1
            changed_count += 1
        else:
            skipped_count += 1
    admin_module.db.session.commit()
    if changed_count:
        # 批量导入修正会改变考勤天数，需同步重算加班查询页数据源
        admin_module._sync_manager_month_stats(month)
    return jsonify(
        admin_module._import_summary(
            success_count, skipped_count, failed_count, changed_count, errors
        )
    )


@admin_required
def download_employee_attendance_override_template():
    from routes import admin_core as admin_module
    month = admin_module._validate_month(request.args.get("month")) or datetime.now().strftime("%Y-%m")
    return admin_module._override_workbook_response(
        admin_module._build_employee_override_export_workbook(month, include_real_rows=False),
        "员工考勤修正导入示例.xlsx",
    )


@admin_required
def export_employee_attendance_overrides():
    from routes import admin_core as admin_module
    month = admin_module._validate_month(request.args.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    return admin_module._override_workbook_response(
        admin_module._build_employee_override_export_workbook(month, include_real_rows=True),
        f"员工考勤修正导出_{month}.xlsx",
    )


@admin_required
def import_employee_attendance_overrides():
    from routes import admin_core as admin_module
    month = admin_module._validate_month(request.form.get("month"))
    if not month:
        return jsonify({"error": "请选择有效月份"}), 400
    account_set = admin_module._account_set_for_month(month)
    locked_error = admin_module._ensure_account_set_unlocked(account_set, "导入员工考勤修正")
    if locked_error:
        return locked_error
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "请选择导入文件"}), 400
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    if not rows:
        return jsonify({"error": "empty file"}), 400
    header_idx, header_map = admin_module._parse_header_row(rows, ["月份", "工号", "姓名", "考勤天数", "备注"])
    required = ["月份", "工号", "姓名", "考勤天数", "工时", "半勤天数", "迟到早退", "备注"]
    missing = [key for key in required if key not in header_map]
    if missing:
        return jsonify({"error": f"缺少列：{', '.join(missing)}"}), 400
    success_count = skipped_count = failed_count = changed_count = 0
    errors: list[str] = []
    for row_index, raw in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        row_month = (
            str(raw[header_map["月份"]]).strip()
            if header_map["月份"] < len(raw) and raw[header_map["月份"]] is not None
            else ""
        )
        emp_no = (
            str(raw[header_map["工号"]]).strip()
            if header_map["工号"] < len(raw) and raw[header_map["工号"]] is not None
            else ""
        )
        if not row_month and not emp_no:
            skipped_count += 1
            continue
        if row_month != month:
            failed_count += 1
            errors.append(f"第 {row_index} 行：月份 {row_month or '空'} 与当前月份 {month} 不一致")
            continue
        employee = admin_module.Employee.query.filter_by(emp_no=emp_no).first()
        if not employee or employee.is_manager:
            failed_count += 1
            errors.append(f"第 {row_index} 行：工号 {emp_no or '空'} 未找到普通员工")
            continue
        updates: dict[str, object] = {}
        numeric_error = False
        for key in ("attendance_days", "work_hours"):
            label = admin_module._EMPLOYEE_OVERRIDE_LABELS[key]
            value = raw[header_map[label]] if header_map[label] < len(raw) else None
            parsed, error = admin_module._nullable_float({key: value}, key)
            if error:
                failed_count += 1
                errors.append(f"第 {row_index} 行：{error}")
                numeric_error = True
                break
            if value not in (None, ""):
                updates[key] = parsed
        if numeric_error:
            continue
        # 实际出勤天数为非必填列：旧导入文件可能不含该列，缺列时跳过不覆盖。
        actual_label = admin_module._EMPLOYEE_OVERRIDE_LABELS["actual_attendance_days"]
        if actual_label in header_map:
            value = raw[header_map[actual_label]] if header_map[actual_label] < len(raw) else None
            parsed, error = admin_module._nullable_float({"actual_attendance_days": value}, "actual_attendance_days")
            if error:
                failed_count += 1
                errors.append(f"第 {row_index} 行：{error}")
                numeric_error = True
            elif value not in (None, ""):
                updates["actual_attendance_days"] = parsed
        if numeric_error:
            continue
        for key in ("half_days", "late_early_minutes"):
            label = admin_module._EMPLOYEE_OVERRIDE_LABELS[key]
            value = raw[header_map[label]] if header_map[label] < len(raw) else None
            parsed, error = admin_module._nullable_int({key: value}, key)
            if error:
                failed_count += 1
                errors.append(f"第 {row_index} 行：{error}")
                numeric_error = True
                break
            if value not in (None, ""):
                updates[key] = parsed
        if numeric_error:
            continue
        remark_value = raw[header_map["备注"]] if header_map["备注"] < len(raw) else None
        if remark_value not in (None, ""):
            updates["remark"] = str(remark_value).strip()
        row_obj = admin_module.EmployeeAttendanceOverride.query.filter_by(
            emp_id=employee.id, month=month
        ).first()
        changed = admin_module._apply_employee_override_updates(
            row_obj, employee.id, month, updates, "import", file.filename
        )
        if changed:
            success_count += 1
            changed_count += 1
        else:
            skipped_count += 1
    admin_module.db.session.commit()
    return jsonify(
        admin_module._import_summary(
            success_count, skipped_count, failed_count, changed_count, errors
        )
    )
