from __future__ import annotations

import os
from collections import defaultdict
from copy import copy
from datetime import date, datetime, time, timedelta
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from flask import jsonify, request, g, send_file
from sqlalchemy.orm import joinedload
import openpyxl

from models import db
from models.employee import Employee
from models.department import Department
from models.daily_record import DailyRecord
from models.overtime import OvertimeRecord
from models.leave import LeaveRecord
from models.annual_leave import AnnualLeave
from models.monthly_report import MonthlyReport
from models.account_set import AccountSet
from models.manager_month_stat import ManagerMonthStat
from models.employee_attendance_override import EmployeeAttendanceOverride
from models.user import EMPLOYEE_PAGE_PERMISSION_KEYS, MANAGER_PAGE_PERMISSION_KEYS, UserEmployeeAssignment, UserDepartmentAssignment
from services.attendance_service import AttendanceService
from services.attendance_source_service import (
    EMPLOYEE_STATS_CONTEXT,
    MANAGER_STATS_CONTEXT,
    attendance_views_by_employee,
)
from services.manager_attendance_service import (
    MANAGER_HEADERS,
    _leave_bucket as _manager_leave_bucket,
    ManagerAttendanceOptions,
    build_manager_rows,
    manager_headers,
    normalize_days,
    rows_as_table,
)
from utils.helpers import overlap_duration_days


MANAGER_ATTENDANCE_TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "export_templates" / "manager_attendance.xlsx"
MANAGER_TEMPLATE_TITLE_ROW = 1
MANAGER_TEMPLATE_HEADER_ROW = 2
MANAGER_TEMPLATE_FIRST_DATA_ROW = 3
MANAGER_TEMPLATE_LAST_DATA_ROW = 105
MANAGER_TEMPLATE_NOTICE_ROW = 106
MANAGER_TEMPLATE_SIGN_ROW = 108
MANAGER_TEMPLATE_DATE_ROW = 109


FINAL_HEADERS = [
    "部门名称",
    "人员编号",
    "人员名称",
    "考勤天数",
    "病假（次数）",
    "工伤（次数）",
    "丧假（次数）",
    "事假（次数）",
    "补休（调休）(次)",
    "婚假（次）",
    "病假时长（天）",
    "工伤时长（天）",
    "丧假时长（天）",
    "事假时长（天）",
    "补休（调休）(天)",
    "婚假（天）",
    "工时",
    "半勤天数",
    "备注",
]

LEAVE_COUNT_HEADERS = {
    "病假（次数）",
    "工伤（次数）",
    "丧假（次数）",
    "事假（次数）",
    "补休（调休）(次)",
    "婚假（次）",
}

LEAVE_DURATION_HEADERS = {
    "病假时长（天）",
    "工伤时长（天）",
    "丧假时长（天）",
    "事假时长（天）",
    "补休（调休）(天)",
    "婚假（天）",
}


def _filter_final_columns(headers: list[str], rows: list[list[object]]) -> tuple[list[str], list[list[object]]]:
    requested = (request.args.get("final_headers") or "").strip()
    if requested:
        wanted = [h.strip() for h in requested.split(",") if h.strip()]
        wanted_set = set(wanted)
        keep_indexes: list[int] = []
        filtered_headers: list[str] = []
        for idx, header in enumerate(headers):
            if header in wanted_set:
                keep_indexes.append(idx)
                filtered_headers.append(header)
        filtered_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indexes] for row in rows]
        return filtered_headers, filtered_rows

    show_leave_counts = request.args.get("show_leave_counts", "").strip() in {"1", "true", "True"}
    show_leave_durations = request.args.get("show_leave_durations", "").strip() in {"1", "true", "True"}

    keep_indexes: list[int] = []
    filtered_headers: list[str] = []
    for idx, header in enumerate(headers):
        if header in LEAVE_COUNT_HEADERS and not show_leave_counts:
            continue
        if header in LEAVE_DURATION_HEADERS and not show_leave_durations:
            continue
        keep_indexes.append(idx)
        filtered_headers.append(header)

    filtered_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indexes] for row in rows]
    return filtered_headers, filtered_rows


def _filter_punch_columns(headers: list[str], rows: list[list[object]]) -> tuple[list[str], list[list[object]]]:
    requested = (request.args.get("punch_headers") or "").strip()
    if not requested:
        return headers, rows
    wanted = [h.strip() for h in requested.split(",") if h.strip()]
    wanted_set = set(wanted)
    keep_indexes: list[int] = []
    filtered_headers: list[str] = []
    for idx, header in enumerate(headers):
        if header in wanted_set:
            keep_indexes.append(idx)
            filtered_headers.append(header)
    filtered_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indexes] for row in rows]
    return filtered_headers, filtered_rows


def _filter_columns(headers: list[str], rows: list[list[object]], param_name: str) -> tuple[list[str], list[list[object]]]:
    requested = (request.args.get(param_name) or "").strip()
    if not requested:
        return headers, rows
    wanted = [h.strip() for h in requested.split(",") if h.strip()]
    wanted_set = set(wanted)
    keep_indexes: list[int] = []
    filtered_headers: list[str] = []
    for idx, header in enumerate(headers):
        if header in wanted_set:
            keep_indexes.append(idx)
            filtered_headers.append(header)
    filtered_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indexes] for row in rows]
    return filtered_headers, filtered_rows


def _normalize_leave_type(value: str | None) -> str:
    text = (value or "").strip()
    if text in {"补休(调休)", "补休（调休）"}:
        return "补休（调休）"
    return text


def _leave_bucket(value: str | None) -> str | None:
    text = _normalize_leave_type(value)
    if not text:
        return None
    if "病假" in text:
        return "病假"
    if "工伤" in text:
        return "工伤"
    if "丧假" in text:
        return "丧假"
    if "事假" in text:
        return "事假"
    if "补休" in text or "调休" in text:
        return "补休（调休）"
    if "婚假" in text:
        return "婚假"
    # Source file may use generic "请假"; align it to personal leave bucket.
    if "请假" in text:
        return "事假"
    return None


def _normalized_leave_days(duration: float | int | None) -> float:
    return normalize_days(duration)


def _leave_days_in_month(record: LeaveRecord, month: str) -> float:
    datetime_range = _month_datetime_range(month)
    if not datetime_range:
        return 0.0
    start_dt, end_dt = datetime_range
    return overlap_duration_days(record.start_time, record.end_time, start_dt, end_dt)


def _month_date_range(month: str) -> tuple[date, date] | None:
    try:
        start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    except ValueError:
        return None
    if start.month == 12:
        end = date(start.year + 1, 1, 1)
    else:
        end = date(start.year, start.month + 1, 1)
    return start, end


def _month_datetime_range(month: str) -> tuple[datetime, datetime] | None:
    bounds = _month_date_range(month)
    if not bounds:
        return None
    start, end = bounds
    return datetime.combine(start, time.min), datetime.combine(end, time.min)


def _format_datetime_value(value: datetime | None) -> str:
    if not value:
        return ""
    return value.strftime("%Y-%m-%d %H:%M")


def _has_punch_record(record) -> bool:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    in_times = [str(x).strip() for x in (record.check_in_times or []) if x is not None and str(x).strip()]
    out_times = [str(x).strip() for x in (record.check_out_times or []) if x is not None and str(x).strip()]
    if in_times or out_times:
        return True

    manager_time_keys = (
        "上班1打卡时间",
        "下班1打卡时间",
        "上班2打卡时间",
        "下班2打卡时间",
        "上班3打卡时间",
        "下班3打卡时间",
        "上班4打卡时间",
        "下班4打卡时间",
    )
    punch_data = str(raw.get("刷卡时间数据") or "").strip()
    if punch_data:
        return True
    return any(str(raw.get(key) or "").strip() for key in manager_time_keys)


def _attendance_day_value(record) -> float:
    if not _has_punch_record(record):
        return 0.0
    actual_hours = record.actual_hours or 0
    if actual_hours <= 0:
        actual_hours, _ = _calc_record_work_hours(record)
    if actual_hours < 2:
        return 0.5
    return 1.0


def _normalize_punch_token(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    m = re.search(r"(\d{1,2}:\d{2})", text)
    if not m:
        return text
    hh, mm = m.group(1).split(":")
    return f"{int(hh):02d}:{mm}"


def _punch_events(record) -> set[str]:
    events: set[str] = set()
    for raw in (record.check_in_times or []):
        token = _normalize_punch_token(raw)
        if token:
            events.add(token)
    for raw in (record.check_out_times or []):
        token = _normalize_punch_token(raw)
        if token:
            events.add(token)
    return events


def _punch_count(record) -> int:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    in_events = {_normalize_punch_token(x) for x in (record.check_in_times or [])}
    in_events = {x for x in in_events if x}
    out_events = {_normalize_punch_token(x) for x in (record.check_out_times or [])}
    out_events = {x for x in out_events if x}

    # Some source rows duplicate the same swipe into both in/out arrays.
    overlap = in_events & out_events
    if overlap:
        in_events -= overlap
        out_events -= overlap

    count = len(in_events) + len(out_events)
    if count:
        return count

    manager_time_keys = (
        "上班1打卡时间",
        "下班1打卡时间",
        "上班2打卡时间",
        "下班2打卡时间",
        "上班3打卡时间",
        "下班3打卡时间",
        "上班4打卡时间",
        "下班4打卡时间",
    )
    manager_events = {
        _normalize_punch_token(raw.get(key))
        for key in manager_time_keys
        if _normalize_punch_token(raw.get(key))
    }
    if manager_events:
        return len(manager_events)

    punch_data = str(raw.get("刷卡时间数据") or "").strip()
    if punch_data:
        return len(re.findall(r"(\d{1,2}:\d{2})", punch_data))
    return 0


def _punch_round_count(record) -> int:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    in_events = {_normalize_punch_token(x) for x in (record.check_in_times or [])}
    in_events = {x for x in in_events if x}
    out_events = {_normalize_punch_token(x) for x in (record.check_out_times or [])}
    out_events = {x for x in out_events if x}

    overlap = in_events & out_events
    if overlap:
        in_events -= overlap
        out_events -= overlap

    # Query page shows "打卡轮次" (e.g. 上午+下午 = 2), not raw swipe points.
    rounds = max(len(in_events), len(out_events))
    if rounds:
        return rounds

    manager_pairs = (
        ("上班1打卡时间", "下班1打卡时间"),
        ("上班2打卡时间", "下班2打卡时间"),
        ("上班3打卡时间", "下班3打卡时间"),
        ("上班4打卡时间", "下班4打卡时间"),
    )
    manager_rounds = 0
    for in_key, out_key in manager_pairs:
        if str(raw.get(in_key) or "").strip() or str(raw.get(out_key) or "").strip():
            manager_rounds += 1
    return manager_rounds


def _format_punch_times(values: list[object] | None) -> str:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in (values or []):
        token = _normalize_punch_token(raw)
        if not token or token in seen:
            continue
        seen.add(token)
        normalized.append(token)
    return " / ".join(normalized)


def _repair_mojibake(text: str) -> str:
    try:
        return text.encode("latin1").decode("gbk")
    except Exception:
        return text


def _stringify_raw_punch_value(value: object) -> str:
    if isinstance(value, list):
        normalized = [str(item).strip() for item in value if str(item).strip()]
        return ",".join(normalized)
    return str(value).strip()


def _extract_raw_punch_data_from_dict(raw: dict) -> str:
    if not isinstance(raw, dict):
        return ""

    direct_keys = {"刷卡时间数据", "原始刷卡数据", "刷卡时间", "打卡记录"}
    for key in direct_keys:
        value = raw.get(key)
        if value is not None and _stringify_raw_punch_value(value):
            return _stringify_raw_punch_value(value)

    for key, value in raw.items():
        if value is None or not _stringify_raw_punch_value(value):
            continue
        repaired = _repair_mojibake(str(key))
        if ("刷卡" in repaired and "时间" in repaired) or ("打卡" in repaired and "记录" in repaired):
            return _stringify_raw_punch_value(value)

    manager_time_keys = (
        "上班1打卡时间",
        "下班1打卡时间",
        "上班2打卡时间",
        "下班2打卡时间",
        "上班3打卡时间",
        "下班3打卡时间",
        "上班4打卡时间",
        "下班4打卡时间",
    )
    manager_times: list[str] = []
    for key in manager_time_keys:
        token = _normalize_punch_token(raw.get(key))
        if token:
            manager_times.append(token)
    if manager_times:
        return ",".join(manager_times)

    return ""


def _extract_raw_punch_data(record) -> str:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    fallback_raw = raw.get("fallback_raw_data") if isinstance(raw.get("fallback_raw_data"), dict) else {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    primary_value = _extract_raw_punch_data_from_dict(raw)
    if primary_value:
        return primary_value
    return _extract_raw_punch_data_from_dict(fallback_raw)


def _raw_punch_count(record) -> int:
    raw = _extract_raw_punch_data(record)
    if raw:
        tokens = re.findall(r"(\d{1,2}:\d{2})", raw)
        if tokens:
            return len(tokens)
    return _punch_round_count(record)


def _parse_punch_dt(value: object, record_date) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    # Full datetime like 2026-03-11 18:00
    m_dt = re.search(r"(\d{4}-\d{2}-\d{2})\s+(\d{1,2}:\d{2})", text)
    if m_dt:
        try:
            dt_text = f"{m_dt.group(1)} {m_dt.group(2)}"
            return datetime.strptime(dt_text, "%Y-%m-%d %H:%M")
        except Exception:
            pass
    # Time only like 07:31, attach record_date
    m_tm = re.search(r"(\d{1,2}):(\d{2})", text)
    if m_tm and record_date:
        try:
            hh = int(m_tm.group(1))
            mm = int(m_tm.group(2))
            return datetime.combine(record_date, datetime.min.time()).replace(hour=hh, minute=mm)
        except Exception:
            return None
    return None


def _resolve_shift_for_record(record):
    if record.employee and record.employee.shift_assignment and record.employee.shift_assignment.shift:
        return record.employee.shift_assignment.shift
    if record.shift:
        return record.shift
    return None


def _parse_slot_dt(record_date, hhmm: str, anchor: datetime | None = None) -> datetime | None:
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", str(hhmm or ""))
    if not m or not record_date:
        return None
    base = datetime.combine(record_date, datetime.min.time()).replace(hour=int(m.group(1)), minute=int(m.group(2)))
    if anchor and base < anchor:
        base += timedelta(days=1)
    return base


def _build_shift_break_windows(record) -> list[tuple[datetime, datetime]]:
    shift = _resolve_shift_for_record(record)
    slots = (shift.time_slots if shift else None) or []
    parsed_slots: list[tuple[datetime, datetime]] = []
    prev_end: datetime | None = None

    for slot in slots:
        if not isinstance(slot, (list, tuple)) or len(slot) != 2:
            continue
        s0 = str(slot[0] or "").strip()
        s1 = str(slot[1] or "").strip()

        # normal pair: ["08:00","11:20"]
        if re.match(r"^\d{1,2}:\d{2}$", s0) and re.match(r"^\d{1,2}:\d{2}$", s1):
            start = _parse_slot_dt(record.record_date, s0, prev_end)
            end = _parse_slot_dt(record.record_date, s1, start)
            if start and end:
                if end < start:
                    end += timedelta(days=1)
                parsed_slots.append((start, end))
                prev_end = end
            continue

        # malformed pair fallback: ["08:00","11:00,12:00-17:00,..."]
        if re.match(r"^\d{1,2}:\d{2}$", s0):
            first_end = s1.split(",")[0].strip()
            if re.match(r"^\d{1,2}:\d{2}$", first_end):
                start = _parse_slot_dt(record.record_date, s0, prev_end)
                end = _parse_slot_dt(record.record_date, first_end, start)
                if start and end:
                    if end < start:
                        end += timedelta(days=1)
                    parsed_slots.append((start, end))
                    prev_end = end

        for a, b in re.findall(r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})", s1):
            start = _parse_slot_dt(record.record_date, a, prev_end)
            end = _parse_slot_dt(record.record_date, b, start)
            if not start or not end:
                continue
            if end < start:
                end += timedelta(days=1)
            parsed_slots.append((start, end))
            prev_end = end

    parsed_slots.sort(key=lambda x: x[0])
    breaks: list[tuple[datetime, datetime]] = []
    for idx in range(len(parsed_slots) - 1):
        curr_end = parsed_slots[idx][1]
        next_start = parsed_slots[idx + 1][0]
        if next_start > curr_end:
            breaks.append((curr_end, next_start))
    return breaks


def _calc_two_punch_hours_with_shift_break(record) -> float | None:
    raw = _extract_raw_punch_data(record)
    times = re.findall(r"(\d{1,2}:\d{2})", raw)
    if len(times) != 2:
        return None
    start = _parse_slot_dt(record.record_date, times[0])
    end = _parse_slot_dt(record.record_date, times[1], start)
    if not start or not end:
        return None
    if end < start:
        end += timedelta(days=1)

    total_seconds = (end - start).total_seconds()
    if total_seconds <= 0:
        return None

    for b_start, b_end in _build_shift_break_windows(record):
        overlap_start = max(start, b_start)
        overlap_end = min(end, b_end)
        if overlap_end > overlap_start:
            total_seconds -= (overlap_end - overlap_start).total_seconds()

    hours = max(total_seconds / 3600.0, 0.0)
    return round(hours, 2)


def _calc_record_work_hours(record) -> tuple[float, int]:
    special_hours = _calc_two_punch_hours_with_shift_break(record)
    if special_hours is not None:
        return special_hours, 0

    in_times: list[datetime] = []
    out_times: list[datetime] = []
    in_seen: set[tuple[int, int, int, int, int]] = set()
    out_seen: set[tuple[int, int, int, int, int]] = set()

    for raw in (record.check_in_times or []):
        dt = _parse_punch_dt(raw, record.record_date)
        if not dt:
            continue
        key = (dt.year, dt.month, dt.day, dt.hour, dt.minute)
        if key in in_seen:
            continue
        in_seen.add(key)
        in_times.append(dt)

    for raw in (record.check_out_times or []):
        dt = _parse_punch_dt(raw, record.record_date)
        if not dt:
            continue
        key = (dt.year, dt.month, dt.day, dt.hour, dt.minute)
        if key in out_seen:
            continue
        out_seen.add(key)
        out_times.append(dt)

    in_times.sort()
    out_times.sort()
    used_out: set[int] = set()
    total_hours = 0.0
    unmatched = 0

    for in_dt in in_times:
        match_idx = None
        match_out = None
        for idx, out_dt in enumerate(out_times):
            if idx in used_out:
                continue
            candidate = out_dt
            if candidate < in_dt:
                candidate = candidate + timedelta(days=1)
            if candidate >= in_dt:
                match_idx = idx
                match_out = candidate
                break
        if match_idx is None or match_out is None:
            unmatched += 1
            continue
        used_out.add(match_idx)
        hours = (match_out - in_dt).total_seconds() / 3600.0
        if 0 < hours <= 20:
            total_hours += hours
        else:
            unmatched += 1

    unmatched += max(len(out_times) - len(used_out), 0)
    if total_hours <= 0 and (record.actual_hours or 0) > 0:
        source_hours = float(record.actual_hours or 0)
        if getattr(record, "source", "") == "manager":
            source_hours = source_hours / 60.0
        return round(source_hours, 2), unmatched
    return round(total_hours, 2), unmatched


def _accessible_emp_ids() -> list[int]:
    if getattr(g, "current_user", None) is None:
        return []
    if g.current_user.role == "admin":
        return [e.id for e in Employee.query.with_entities(Employee.id).all()]
        
    emp_rows = UserEmployeeAssignment.query.filter_by(user_id=g.current_user.id).all()
    dept_rows = UserDepartmentAssignment.query.filter_by(user_id=g.current_user.id).all()
        
    ids = {r.emp_id for r in emp_rows}
    assigned_dept_ids = {r.dept_id for r in dept_rows}
    
    if assigned_dept_ids:
        # Find all descendant departments
        all_departments = Department.query.with_entities(Department.id, Department.parent_id).all()
        children_map = {}
        for d in all_departments:
            children_map.setdefault(d.parent_id, []).append(d.id)
            
        expanded_dept_ids = set(assigned_dept_ids)
        queue = list(assigned_dept_ids)
        while queue:
            curr = queue.pop(0)
            for child_id in children_map.get(curr, []):
                if child_id not in expanded_dept_ids:
                    expanded_dept_ids.add(child_id)
                    queue.append(child_id)
                    
        dept_emp_ids = Employee.query.with_entities(Employee.id).filter(Employee.dept_id.in_(expanded_dept_ids)).all()
        ids.update(row.id for row in dept_emp_ids)
        
    return list(ids)


def _can_access_query_center() -> bool:
    return g.current_user.role == "admin" or g.current_user.has_any_page_access(
        (*MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)
    )


def _non_manager_emp_ids(emp_ids: list[int]) -> list[int]:
    if not emp_ids:
        return []
    rows = (
        Employee.query.with_entities(Employee.id)
        .filter(Employee.id.in_(emp_ids), Employee.is_manager.is_(False))
        .all()
    )
    allowed = {row.id for row in rows}
    return [emp_id for emp_id in emp_ids if emp_id in allowed]


def _manager_emp_ids(emp_ids: list[int]) -> list[int]:
    if not emp_ids:
        return []
    rows = (
        Employee.query.with_entities(Employee.id)
        .filter(Employee.id.in_(emp_ids), Employee.is_manager.is_(True))
        .all()
    )
    allowed = {row.id for row in rows}
    return [emp_id for emp_id in emp_ids if emp_id in allowed]


def _accessible_manager_emp_ids() -> list[int]:
    manager_ids = set(_manager_emp_ids(_accessible_emp_ids()))
    profile_emp_no = (g.current_user.profile_emp_no or "").strip()
    if profile_emp_no:
        profile_manager = Employee.query.with_entities(Employee.id).filter_by(emp_no=profile_emp_no, is_manager=True).first()
        if profile_manager:
            manager_ids.add(profile_manager.id)
    return list(manager_ids)


def _manager_factory_rest_days(account_set: AccountSet | None) -> float:
    if account_set is None:
        return 0.0
    if not account_set.factory_rest_entries:
        return 0.0
    total = 0.0
    for item in account_set.factory_rest_entries:
        if item.rest_period == "full":
            total += 1.0
        elif item.rest_period in {"am", "pm"}:
            total += 0.5
    return total


def _manager_factory_rest_requires_detail(account_set: AccountSet | None) -> bool:
    if account_set is None:
        return False
    return not account_set.factory_rest_entries and float(account_set.factory_rest_days or 0) > 0


def _manager_options() -> ManagerAttendanceOptions:
    month = _resolve_query_month()
    account_set = AccountSet.query.filter_by(month=month).first()
    return ManagerAttendanceOptions(
        month=month,
        factory_rest_days=_manager_factory_rest_days(account_set),
        monthly_benefit_days=(account_set.monthly_benefit_days if account_set else 0.0) or 0.0,
    )


def _pick_emp_id() -> int | None:
    requested = request.args.get("emp_id", type=int)
    allowed = _non_manager_emp_ids(_accessible_emp_ids())
    if requested and requested in allowed:
        return requested
    return allowed[0] if allowed else None


def _requested_emp_ids() -> list[int]:
    ids: list[int] = []
    for raw in request.args.getlist("emp_ids"):
        for part in str(raw).split(","):
            text = part.strip()
            if text.isdigit():
                ids.append(int(text))
    if ids:
        return ids

    single = request.args.get("emp_id", type=int)
    return [single] if single else []


def _keyword_filtered_emp_ids(base_ids: list[int]) -> list[int]:
    keyword = (request.args.get("emp_keyword") or "").strip()
    if not keyword:
        return base_ids
    if not base_ids:
        return []

    like_kw = f"%{keyword}%"
    rows = (
        Employee.query.with_entities(Employee.id)
        .filter(Employee.id.in_(base_ids))
        .filter((Employee.emp_no.like(like_kw)) | (Employee.name.like(like_kw)))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    matched = [row.id for row in rows]
    if not matched:
        return []
    matched_set = set(matched)
    # keep base_ids order
    return [emp_id for emp_id in base_ids if emp_id in matched_set]


def _pick_emp_ids() -> list[int]:
    requested = _requested_emp_ids()
    allowed = _non_manager_emp_ids(_accessible_emp_ids())
    if requested:
        allowed_set = set(allowed)
        filtered = [emp_id for emp_id in requested if emp_id in allowed_set]
        # keep order while deduplicating
        return _keyword_filtered_emp_ids(list(dict.fromkeys(filtered)))
    return _keyword_filtered_emp_ids(allowed)


def _resolve_query_month() -> str:
    active_set = AccountSet.query.filter_by(is_active=True).first()
    return request.args.get("month") or (active_set.month if active_set else datetime.now().strftime("%Y-%m"))


def _build_final_rows(month: str, emp_ids: list[int]) -> list[list[object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    rows: list[list[object]] = []
    date_range = _month_date_range(month)
    datetime_range = _month_datetime_range(month)
    overrides = {
        row.emp_id: row
        for row in EmployeeAttendanceOverride.query.filter(
            EmployeeAttendanceOverride.month == month,
            EmployeeAttendanceOverride.emp_id.in_(emp_ids),
        ).all()
    }

    daily_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT) if date_range else {}

    leave_by_emp: dict[int, list[LeaveRecord]] = defaultdict(list)
    if datetime_range:
        start_dt, end_dt = datetime_range
        leave_records = (
            LeaveRecord.query.filter(LeaveRecord.emp_id.in_(emp_ids))
            .filter(LeaveRecord.start_time < end_dt, LeaveRecord.end_time > start_dt)
            .all()
        )
        for record in leave_records:
            leave_by_emp[record.emp_id].append(record)

    for employee in employees:
        daily_rows = daily_by_emp.get(employee.id, [])
        leave_rows = leave_by_emp.get(employee.id, [])

        leave_count = {"病假": 0, "工伤": 0, "丧假": 0, "事假": 0, "补休（调休）": 0, "婚假": 0}
        leave_days = {"病假": 0.0, "工伤": 0.0, "丧假": 0.0, "事假": 0.0, "补休（调休）": 0.0, "婚假": 0.0}
        for row in leave_rows:
            leave_type = _leave_bucket(row.leave_type)
            if not leave_type:
                continue
            leave_count[leave_type] += 1
            leave_days[leave_type] += _normalized_leave_days(_leave_days_in_month(row, month))

        day_work_stats = [_calc_record_work_hours(r) for r in daily_rows]
        attendance_days = round(sum(_attendance_day_value(r) for r in daily_rows), 2)
        half_days = sum(
            1
            for idx, r in enumerate(daily_rows)
            if _punch_count(r) == 2 and 2 <= (day_work_stats[idx][0]) < 5.1
        )
        work_hours = round(sum(x[0] for x in day_work_stats), 2)
        override = overrides.get(employee.id)
        if override:
            if override.attendance_days is not None:
                attendance_days = round(float(override.attendance_days or 0), 2)
            if override.work_hours is not None:
                work_hours = round(float(override.work_hours or 0), 2)
            if override.half_days is not None:
                half_days = int(override.half_days or 0)

        row = [
            employee.department.dept_name if employee.department else "",
            employee.emp_no,
            employee.name,
            attendance_days,
            leave_count["病假"],
            leave_count["工伤"],
            leave_count["丧假"],
            leave_count["事假"],
            leave_count["补休（调休）"],
            leave_count["婚假"],
            round(leave_days["病假"], 2),
            round(leave_days["工伤"], 2),
            round(leave_days["丧假"], 2),
            round(leave_days["事假"], 2),
            round(leave_days["补休（调休）"], 2),
            round(leave_days["婚假"], 2),
            work_hours,
            half_days,
            "",
        ]
        rows.append(row)

    return rows


def _build_abnormal_rows(month: str, emp_ids: list[int]) -> list[dict[str, object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    data: list[dict[str, object]] = []
    date_range = _month_date_range(month)
    daily_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT) if date_range else {}

    for employee in employees:
        daily_rows = daily_by_emp.get(employee.id, [])
        abnormal_dates = {
            r.record_date.isoformat()
            for r in daily_rows
            if r.record_date and _raw_punch_count(r) in {1, 3}
        }
        data.append(
            {
                "dept_name": employee.department.dept_name if employee.department else "",
                "emp_no": employee.emp_no,
                "name": employee.name,
                "abnormal_count": len(abnormal_dates),
            }
        )
    return data


def _build_leave_detail_rows(month: str, emp_ids: list[int], leave_type: str | None) -> list[dict[str, object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    employee_by_id = {employee.id: employee for employee in employees}
    datetime_range = _month_datetime_range(month)
    if not datetime_range or not employees:
        return []

    start_dt, end_dt = datetime_range
    normalized_target = _normalize_leave_type(leave_type)
    rows: list[dict[str, object]] = []
    leave_records = (
        LeaveRecord.query.filter(LeaveRecord.emp_id.in_(emp_ids))
        .filter(LeaveRecord.start_time < end_dt, LeaveRecord.end_time > start_dt)
        .order_by(LeaveRecord.start_time.asc(), LeaveRecord.id.asc())
        .all()
    )
    for record in leave_records:
        bucket = _leave_bucket(record.leave_type)
        if normalized_target and bucket != normalized_target:
            continue
        employee = employee_by_id.get(record.emp_id)
        if not employee:
            continue
        rows.append(
            {
                "dept_name": employee.department.dept_name if employee.department else "",
                "name": employee.name,
                "leave_type": bucket or _normalize_leave_type(record.leave_type),
                "start_time": _format_datetime_value(record.start_time),
                "end_time": _format_datetime_value(record.end_time),
                "duration": round(_normalized_leave_days(_leave_days_in_month(record, month)), 2),
                "reason": record.reason or "",
            }
        )
    return rows


def _build_department_hours_rows(month: str, emp_ids: list[int]) -> list[dict[str, object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    totals: dict[str, float] = {}
    dept_member_counts: dict[str, int] = {}
    overrides = {
        row.emp_id: row
        for row in EmployeeAttendanceOverride.query.filter(
            EmployeeAttendanceOverride.month == month,
            EmployeeAttendanceOverride.emp_id.in_(emp_ids),
        ).all()
    }

    for employee in employees:
        dept_name = employee.department.dept_name if employee.department else "未分配部门"
        totals.setdefault(dept_name, 0.0)
        dept_member_counts[dept_name] = dept_member_counts.get(dept_name, 0) + 1

    if not employees:
        return []
    rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
    for employee in employees:
        dept_name = employee.department.dept_name if employee.department else "未分配部门"
        work_hours = round(sum(_calc_record_work_hours(row)[0] for row in rows_by_emp.get(employee.id, [])), 2)
        override = overrides.get(employee.id)
        if override and override.work_hours is not None:
            work_hours = round(float(override.work_hours or 0), 2)
        totals.setdefault(dept_name, 0.0)
        totals[dept_name] += work_hours

    return [
        {
            "dept_name": k,
            "total_hours": round(v, 2),
            "member_count": dept_member_counts.get(k, 0)
        }
        for k, v in sorted(totals.items(), key=lambda x: x[0])
    ]


def _top_level_department_name(department: Department | None) -> str:
    if not department:
        return "未分配部门"
    cursor = department
    visited: set[int] = set()
    while cursor.parent and cursor.id not in visited:
        visited.add(cursor.id)
        cursor = cursor.parent
    return cursor.dept_name or "未分配部门"


def _build_manager_department_hours_rows(month: str, emp_ids: list[int]) -> list[dict[str, object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    totals: dict[str, float] = {}
    dept_member_counts: dict[str, int] = {}

    for employee in employees:
        dept_name = _top_level_department_name(employee.department)
        totals.setdefault(dept_name, 0.0)
        dept_member_counts[dept_name] = dept_member_counts.get(dept_name, 0) + 1

    if not employees:
        return []
    rows_by_emp = attendance_views_by_employee(month, employees, MANAGER_STATS_CONTEXT)
    for employee in employees:
        dept_name = _top_level_department_name(employee.department)
        for row in rows_by_emp.get(employee.id, []):
            totals.setdefault(dept_name, 0.0)
            if row.source == "manager":
                totals[dept_name] += float(row.actual_hours or 0) / 60.0
            else:
                totals[dept_name] += float(row.actual_hours or 0)

    return [
        {
            "dept_name": k,
            "total_hours": round(v, 2),
            "member_count": dept_member_counts.get(k, 0)
        }
        for k, v in sorted(totals.items(), key=lambda x: x[0])
    ]


def account_sets_api():
    if not _can_access_query_center():
        return jsonify({"error": "Forbidden"}), 403
    rows = AccountSet.query.order_by(AccountSet.month.desc()).all()
    return jsonify(
        [
            {
                "id": r.id,
                "month": r.month,
                "name": r.name,
                "is_active": r.is_active,
                "factory_rest_days": _manager_factory_rest_days(r),
                "factory_rest_requires_detail": _manager_factory_rest_requires_detail(r),
                "legacy_factory_rest_days": (r.factory_rest_days or 0),
                "monthly_benefit_days": r.monthly_benefit_days or 0,
            }
            for r in rows
        ]
    )


def departments_api():
    if not _can_access_query_center():
        return jsonify({"error": "Forbidden"}), 403

    emp_ids = _accessible_emp_ids()
    if not emp_ids:
        return jsonify([])

    dept_ids = {
        row.dept_id
        for row in Employee.query.with_entities(Employee.dept_id)
        .filter(Employee.id.in_(emp_ids), Employee.dept_id.isnot(None))
        .all()
    }
    if not dept_ids:
        return jsonify([])

    all_ids = set(dept_ids)
    cursor_ids = set(dept_ids)
    while cursor_ids:
        parents = (
            Department.query.with_entities(Department.parent_id)
            .filter(Department.id.in_(cursor_ids), Department.parent_id.isnot(None))
            .all()
        )
        next_ids = {row.parent_id for row in parents if row.parent_id and row.parent_id not in all_ids}
        all_ids.update(next_ids)
        cursor_ids = next_ids

    depts = Department.query.filter(Department.id.in_(all_ids)).order_by(Department.dept_name.asc()).all()
    return jsonify(
        [{"id": d.id, "dept_no": d.dept_no, "dept_name": d.dept_name, "parent_id": d.parent_id} for d in depts]
    )


def home_manager_summary_api():
    if not g.current_user.can_access_page("query_home"):
        return jsonify({"error": "Forbidden"}), 403

    account_sets = AccountSet.query.order_by(AccountSet.month.desc()).all()
    if not account_sets:
        return jsonify({"has_data": False, "empty_state": "暂无账套，暂无数据", "month": "", "account_set_name": ""})

    month = (request.args.get("month") or "").strip()
    if not month:
        active_set = next((row for row in account_sets if row.is_active), None)
        month = active_set.month if active_set else account_sets[0].month

    account_set = next((row for row in account_sets if row.month == month), None)
    if not account_set:
        return jsonify({"has_data": False, "empty_state": "所选账套无数据", "month": month, "account_set_name": ""})

    profile_emp_no = (g.current_user.profile_emp_no or "").strip()
    if not profile_emp_no:
        return jsonify(
            {
                "has_data": False,
                "empty_state": "账号未绑定管理人员工号，暂无数据",
                "month": month,
                "account_set_name": account_set.name,
            }
        )

    manager = Employee.query.options(joinedload(Employee.department)).filter_by(emp_no=profile_emp_no).first()
    if not manager or not manager.is_manager:
        return jsonify(
            {
                "has_data": False,
                "empty_state": "未找到匹配的管理人员数据",
                "month": month,
                "account_set_name": account_set.name,
            }
        )

    rows = build_manager_rows(
        ManagerAttendanceOptions(
            month=month,
            factory_rest_days=_manager_factory_rest_days(account_set),
            monthly_benefit_days=(account_set.monthly_benefit_days or 0.0),
        ),
        [manager.id],
    )
    if not rows:
        return jsonify(
            {
                "has_data": False,
                "empty_state": "当前账套暂无管理人员考勤数据",
                "month": month,
                "account_set_name": account_set.name,
            }
        )

    summary = rows[0]
    year = int(month.split("-", 1)[0])
    month_no = int(month.split("-", 1)[1])
    overtime_row = ManagerMonthStat.query.filter_by(emp_id=manager.id, year=year, stat_type="overtime").first()
    if overtime_row:
        overtime_remaining = float(overtime_row.prev_dec or 0)
        overtime_remaining += sum(float(getattr(overtime_row, f"m{index}") or 0) for index in range(1, month_no + 1))
        overtime_remaining = round(overtime_remaining, 2)
    else:
        overtime_remaining = 0.0

    annual_leave_row = ManagerMonthStat.query.filter_by(emp_id=manager.id, year=year, stat_type="annual_leave").first()
    if annual_leave_row:
        used_benefit_total = sum(float(getattr(annual_leave_row, f"m{index}") or 0) for index in range(1, month_no + 1))
        benefit_remaining = round(12.0 - used_benefit_total, 2)
    else:
        benefit_remaining = 12.0

    factory_rest_entries = sorted(
        account_set.factory_rest_entries,
        key=lambda item: item.rest_date.isoformat() if item.rest_date else "",
    )
    serialized_factory_rest_entries = [
        {
            "date": item.rest_date.isoformat() if item.rest_date else None,
            "period": item.rest_period,
        }
        for item in factory_rest_entries
    ]

    return jsonify(
        {
            "has_data": True,
            "empty_state": "",
            "month": month,
            "account_set_name": account_set.name,
            "manager": {
                "emp_no": manager.emp_no,
                "name": manager.name,
                "dept_name": manager.department.dept_name if manager.department else "",
            },
            "summary": {
                "attendance_days": summary.get("attendance_days", 0),
                "personal_sick_days": summary.get("personal_sick_days", 0),
                "injury_days": summary.get("injury_days", 0),
                "business_trip_days": summary.get("business_trip_days", 0),
                "marriage_days": summary.get("marriage_days", 0),
                "funeral_days": summary.get("funeral_days", 0),
                "late_early_minutes": summary.get("late_early_minutes", 0),
                "benefit_days": benefit_remaining,
                "overtime_remaining_days": round(overtime_remaining, 2),
            },
            "factory_rest_entries": serialized_factory_rest_entries,
            "support_message": "如对考勤数据有疑问，请联系信息中心协助核对处理。",
        }
    )



def punch_records_api():
    emp_ids = _pick_emp_ids()
    dept_id = request.args.get("dept_id", type=int)
    if dept_id:
        dept_emp_ids = {
            e.id for e in Employee.query.with_entities(Employee.id).filter(Employee.dept_id == dept_id).all()
        }
        emp_ids = [x for x in emp_ids if x in dept_emp_ids]
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    employees = Employee.query.options(joinedload(Employee.department)).filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all()
    rows = []
    rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
    for employee in employees:
        rows.extend(rows_by_emp.get(employee.id, []))
    rows.sort(key=lambda row: ((row.employee.emp_no if row.employee else ""), row.record_date or date.min), reverse=True)

    return jsonify(
        [
            {
                "date": r.record_date.isoformat() if r.record_date else "",
                "emp_no": r.employee.emp_no if r.employee else "",
                "name": r.employee.name if r.employee else "",
                "dept_name": r.employee.department.dept_name if r.employee and r.employee.department else "",
                "raw_punch_data": _extract_raw_punch_data(r),
                "check_in_times": _format_punch_times(r.check_in_times),
                "check_out_times": _format_punch_times(r.check_out_times),
                "punch_count": _raw_punch_count(r),
                "actual_hours": _calc_record_work_hours(r)[0],
                "late_minutes": r.late_minutes or 0,
                "early_leave_minutes": r.early_leave_minutes or 0,
                "exception_reason": r.exception_reason or "",
            }
            for r in rows
        ]
    )


def punch_records_export_api():
    emp_ids = _pick_emp_ids()
    dept_id = request.args.get("dept_id", type=int)
    if dept_id:
        dept_emp_ids = {
            e.id for e in Employee.query.with_entities(Employee.id).filter(Employee.dept_id == dept_id).all()
        }
        emp_ids = [x for x in emp_ids if x in dept_emp_ids]
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    employees = Employee.query.options(joinedload(Employee.department)).filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all()
    rows = []
    rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
    for employee in employees:
        rows.extend(rows_by_emp.get(employee.id, []))
    rows.sort(key=lambda row: ((row.employee.emp_no if row.employee else ""), row.record_date or date.min), reverse=True)

    punch_headers = [
        "日期",
        "员工编号",
        "员工姓名",
        "部门",
        "原始打卡数据",
        "上班打卡",
        "下班打卡",
        "打卡次数",
        "实出勤小时",
        "迟到分钟",
        "早退分钟",
        "异常原因",
    ]

    row_data = []
    for r in rows:
        row_data.append(
            [
                r.record_date.isoformat() if r.record_date else "",
                r.employee.emp_no if r.employee else "",
                r.employee.name if r.employee else "",
                r.employee.department.dept_name if r.employee and r.employee.department else "",
                _extract_raw_punch_data(r),
                _format_punch_times(r.check_in_times),
                _format_punch_times(r.check_out_times),
                _raw_punch_count(r),
                _calc_record_work_hours(r)[0],
                r.late_minutes or 0,
                r.early_leave_minutes or 0,
                r.exception_reason or "",
            ]
        )

    headers, row_data = _filter_punch_columns(punch_headers, row_data)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "打卡数据查询"
    ws.append(headers)
    for row in row_data:
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"打卡数据查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def punch_records_modal_export_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "原始刷卡记录"
    ws.append(["部门", "姓名", "日期", "原始打卡数据"])

    for employee in employees:
        for row in rows_by_emp.get(employee.id, []):
            ws.append(
                [
                    employee.department.dept_name if employee.department else "",
                    employee.name,
                    row.record_date.isoformat() if row.record_date else "",
                    _extract_raw_punch_data(row),
                ]
            )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"原始刷卡记录_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def manager_punch_records_api():
    emp_ids = _accessible_manager_emp_ids()
    requested_ids = _requested_emp_ids()
    if requested_ids:
        allowed = set(emp_ids)
        emp_ids = [emp_id for emp_id in requested_ids if emp_id in allowed]
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    rows = []
    rows_by_emp = attendance_views_by_employee(month, employees, MANAGER_STATS_CONTEXT)
    for employee in employees:
        rows.extend(rows_by_emp.get(employee.id, []))
    rows.sort(key=lambda row: ((row.employee.emp_no if row.employee else ""), row.record_date or date.min), reverse=True)

    return jsonify(
        [
            {
                "date": r.record_date.isoformat() if r.record_date else "",
                "name": r.employee.name if r.employee else "",
                "dept_name": r.employee.department.dept_name if r.employee and r.employee.department else "",
                "raw_punch_data": _extract_raw_punch_data(r),
                "late_minutes": r.late_minutes or 0,
                "early_leave_minutes": r.early_leave_minutes or 0,
            }
            for r in rows
        ]
    )


def leave_records_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    leave_type = request.args.get("leave_type")
    return jsonify(_build_leave_detail_rows(month, emp_ids, leave_type))


def manager_leave_records_api():
    emp_ids = _accessible_manager_emp_ids()
    requested_ids = _requested_emp_ids()
    if requested_ids:
        allowed = set(emp_ids)
        emp_ids = [emp_id for emp_id in requested_ids if emp_id in allowed]
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    leave_bucket = request.args.get("leave_bucket")
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    employee_by_id = {employee.id: employee for employee in employees}
    datetime_range = _month_datetime_range(month)
    if not datetime_range:
        return jsonify([])

    start_dt, end_dt = datetime_range
    rows = []
    leave_records = (
        LeaveRecord.query.filter(LeaveRecord.emp_id.in_(emp_ids))
        .filter(LeaveRecord.start_time < end_dt, LeaveRecord.end_time > start_dt)
        .order_by(LeaveRecord.start_time.asc(), LeaveRecord.id.asc())
        .all()
    )
    for record in leave_records:
        bucket = _manager_leave_bucket(record.leave_type)
        if leave_bucket and bucket != leave_bucket:
            continue
        employee = employee_by_id.get(record.emp_id)
        if not employee:
            continue
        rows.append(
            {
                "dept_name": employee.department.dept_name if employee.department else "",
                "name": employee.name,
                "leave_type": _normalize_leave_type(record.leave_type),
                "start_time": _format_datetime_value(record.start_time),
                "end_time": _format_datetime_value(record.end_time),
                "duration": round(normalize_days(_leave_days_in_month(record, month)), 2),
                "reason": record.reason or "",
            }
        )
    return jsonify(rows)


def leave_records_export_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    leave_type = request.args.get("leave_type")
    rows = _build_leave_detail_rows(month, emp_ids, leave_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "请假明细"
    ws.append(["部门", "姓名", "请假类型", "开始时间", "结束时间", "时长", "事由"])
    for row in rows:
        ws.append(
            [
                row["dept_name"],
                row["name"],
                row["leave_type"],
                row["start_time"],
                row["end_time"],
                row["duration"],
                row["reason"],
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    export_name = _normalize_leave_type(leave_type) or "请假"
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{export_name}明细_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def summary_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify({"error": "No employee assigned"}), 404

    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    year = request.args.get("year", type=int) or datetime.now().year

    monthly = AttendanceService.monthly_summary(emp_id, month)
    yearly = AttendanceService.yearly_summary(emp_id, year)
    deduction = AttendanceService.deduction_calc(emp_id, month)
    annual = AttendanceService.annual_leave_balance(emp_id, year)

    return jsonify(
        {
            "emp_id": emp_id,
            "month": month,
            "year": year,
            "monthly": monthly,
            "yearly": yearly,
            "deduction": deduction,
            "annual_leave": annual,
        }
    )


def daily_records_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify([])

    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    employee = db.session.get(Employee, emp_id)
    rows = attendance_views_by_employee(month, [employee], EMPLOYEE_STATS_CONTEXT).get(emp_id, []) if employee else []
    rows.sort(key=lambda row: row.record_date or date.min, reverse=True)
    return jsonify(
        [
            {
                "date": r.record_date.isoformat(),
                "expected_hours": r.expected_hours,
                "actual_hours": r.actual_hours,
                "absent_hours": r.absent_hours,
                "leave_hours": r.leave_hours,
                "leave_type": r.leave_type,
                "overtime_hours": r.overtime_hours,
                "overtime_type": r.overtime_type,
                "late_minutes": r.late_minutes,
                "early_leave_minutes": r.early_leave_minutes,
                "exception_reason": r.exception_reason,
            }
            for r in rows
        ]
    )


def overtime_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify([])

    rows = OvertimeRecord.query.filter_by(emp_id=emp_id).order_by(OvertimeRecord.start_time.desc()).all()
    return jsonify(
        [
            {
                "overtime_no": r.overtime_no,
                "start_time": r.start_time.isoformat() if r.start_time else None,
                "end_time": r.end_time.isoformat() if r.end_time else None,
                "effective_hours": r.effective_hours,
                "is_weekend": r.is_weekend,
                "is_holiday": r.is_holiday,
                "salary_option": r.salary_option,
                "reason": r.reason,
                "approval_status": r.approval_status,
            }
            for r in rows
        ]
    )


def leave_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify([])

    rows = LeaveRecord.query.filter_by(emp_id=emp_id).order_by(LeaveRecord.start_time.desc()).all()
    return jsonify(
        [
            {
                "leave_no": r.leave_no,
                "apply_date": r.apply_date.isoformat() if r.apply_date else None,
                "leave_type": r.leave_type,
                "start_time": r.start_time.isoformat() if r.start_time else None,
                "end_time": r.end_time.isoformat() if r.end_time else None,
                "duration": r.duration,
                "reason": r.reason,
                "approval_status": r.approval_status,
            }
            for r in rows
        ]
    )


def annual_leave_api():
    emp_id = _pick_emp_id()
    year = request.args.get("year", type=int) or datetime.now().year
    if not emp_id:
        return jsonify({"year": year, "total_days": 0, "used_days": 0, "remaining_days": 0})

    row = AnnualLeave.query.filter_by(emp_id=emp_id, year=year).first()
    if not row:
        return jsonify({"year": year, "total_days": 0, "used_days": 0, "remaining_days": 0})
    return jsonify(
        {
            "year": row.year,
            "total_days": row.total_days,
            "used_days": row.used_days,
            "remaining_days": row.remaining_days,
        }
    )


def final_data_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"headers": [], "rows": [], "error": "No employee assigned"})

    month = _resolve_query_month()
    requested_ids = _requested_emp_ids()
    rows = _build_final_rows(month, emp_ids)
    headers, rows = _filter_final_columns(FINAL_HEADERS, rows)

    return jsonify(
        {
            "headers": headers,
            "rows": rows,
            "month": month,
            "emp_ids": emp_ids,
            "mode": "all" if not requested_ids else ("single" if len(emp_ids) == 1 else "multi"),
        }
    )


def final_data_export_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    rows = _build_final_rows(month, emp_ids)
    headers, rows = _filter_final_columns(FINAL_HEADERS, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤数据查询"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"考勤数据查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def summary_download_export_api():
    requested = _requested_emp_ids()
    accessible = _accessible_emp_ids()
    if requested:
        accessible_set = set(accessible)
        full_emp_ids = [emp_id for emp_id in requested if emp_id in accessible_set]
        full_emp_ids = _keyword_filtered_emp_ids(list(dict.fromkeys(full_emp_ids)))
    else:
        full_emp_ids = _keyword_filtered_emp_ids(accessible)

    if not full_emp_ids:
        return jsonify({"error": "No employee selected"}), 400

    month = _resolve_query_month()
    year = int(month.split("-", 1)[0]) if month else datetime.now().year
    sheets_raw = (request.args.get("sheets") or "final,punch").strip()
    sheets_list = [s.strip() for s in sheets_raw.split(",") if s.strip()]

    emp_ids = _non_manager_emp_ids(full_emp_ids)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. final：员工考勤数据查询
    if "final" in sheets_list:
        final_rows = _build_final_rows(month, emp_ids)
        headers, rows = _filter_final_columns(FINAL_HEADERS, final_rows)
        ws = wb.create_sheet("考勤数据查询")
        ws.append(headers)
        for row in rows:
            ws.append(row)

    # 2. punch：员工打卡数据查询
    if "punch" in sheets_list:
        employees = Employee.query.options(joinedload(Employee.department)).filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all()
        punch_rows = []
        rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
        for employee in employees:
            punch_rows.extend(rows_by_emp.get(employee.id, []))
        punch_rows.sort(key=lambda row: ((row.employee.emp_no if row.employee else ""), row.record_date or date.min), reverse=True)

        punch_headers = [
            "日期", "员工编号", "员工姓名", "部门", "原始打卡数据",
            "上班打卡", "下班打卡", "打卡次数", "实出勤小时",
            "迟到分钟", "早退分钟", "异常原因",
        ]

        punch_row_data = []
        for r in punch_rows:
            punch_row_data.append([
                r.record_date.isoformat() if r.record_date else "",
                r.employee.emp_no if r.employee else "",
                r.employee.name if r.employee else "",
                r.employee.department.dept_name if r.employee and r.employee.department else "",
                _extract_raw_punch_data(r),
                _format_punch_times(r.check_in_times),
                _format_punch_times(r.check_out_times),
                _raw_punch_count(r),
                _calc_record_work_hours(r)[0],
                r.late_minutes or 0,
                r.early_leave_minutes or 0,
                r.exception_reason or "",
            ])

        punch_headers, punch_row_data = _filter_punch_columns(punch_headers, punch_row_data)

        ws2 = wb.create_sheet("打卡数据查询")
        ws2.append(punch_headers)
        for row in punch_row_data:
            ws2.append(row)

    # 3. abnormal：员工异常查询
    if "abnormal" in sheets_list:
        abnormal_rows = _build_abnormal_rows(month, emp_ids)
        headers = ["部门名称", "人员编号", "人员姓名", "异常考勤次数"]
        row_data = []
        for r in abnormal_rows:
            row_data.append([
                r.get("dept_name", ""),
                r.get("emp_no", ""),
                r.get("name", ""),
                r.get("abnormal_count", 0)
            ])
        headers, row_data = _filter_columns(headers, row_data, "abnormal_headers")
        ws3 = wb.create_sheet("员工异常查询")
        ws3.append(headers)
        for row in row_data:
            ws3.append(row)

    # 4. emp_dept_hours：员工部门工时查询
    if "emp_dept_hours" in sheets_list:
        non_mgr_ids = _non_manager_emp_ids(emp_ids)
        dept_hours_rows = _build_department_hours_rows(month, non_mgr_ids)
        headers = ["部门名称", "总工时（小时）", "部门人数"]
        row_data = []
        for r in dept_hours_rows:
            row_data.append([
                r.get("dept_name", ""),
                r.get("total_hours", 0),
                r.get("member_count", 0)
            ])
        headers, row_data = _filter_columns(headers, row_data, "emp_dept_hours_headers")
        ws4 = wb.create_sheet("员工部门工时查询")
        ws4.append(headers)
        for row in row_data:
            ws4.append(row)

    # 5. mgr_attendance：管理人员考勤查询
    if "mgr_attendance" in sheets_list:
        mgr_ids = _manager_emp_ids(full_emp_ids)
        options = ManagerAttendanceOptions(month=month)
        rows = _manager_export_rows_with_top_level_departments(build_manager_rows(options, mgr_ids))
        headers = manager_headers(True)
        row_data = rows_as_table(rows, True)
        headers, row_data = _filter_columns(headers, row_data, "mgr_attendance_headers")
        ws5 = wb.create_sheet("管理人员考勤查询")
        ws5.append(headers)
        for row in row_data:
            ws5.append(row)

    # 6. mgr_overtime：管理人员加班查询
    if "mgr_overtime" in sheets_list:
        from routes.admin_core import _manager_month_rows, _manager_overtime_values
        values = _manager_overtime_values(year)
        mgr_ids = set(_manager_emp_ids(full_emp_ids))
        values = {name: r for name, r in values.items() if r.get("emp_id") in mgr_ids}
        headers = ["部门", "姓名", "前年累积天数", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余调休天数", "备注"]
        row_data = _manager_month_rows(values, "剩余调休天数")
        headers, row_data = _filter_columns(headers, row_data, "mgr_overtime_headers")
        ws6 = wb.create_sheet("管理人员加班查询")
        ws6.append(headers)
        for row in row_data:
            ws6.append(row)

    # 7. mgr_annual_leave：管理人员年假查询
    if "mgr_annual_leave" in sheets_list:
        from routes.admin_core import _annual_leave_value_keys, _manager_annual_leave_values, _manager_month_rows
        values = _manager_annual_leave_values(year)
        mgr_ids = set(_manager_emp_ids(full_emp_ids))
        values = {name: r for name, r in values.items() if r.get("emp_id") in mgr_ids}
        headers = ["部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"]
        row_data = _manager_month_rows(values, "剩余年休天数", _annual_leave_value_keys())
        headers, row_data = _filter_columns(headers, row_data, "mgr_annual_leave_headers")
        ws7 = wb.create_sheet("管理人员年假查询")
        ws7.append(headers)
        for row in row_data:
            ws7.append(row)

    # 8. mgr_dept_hours：管理人员部门工时查询
    if "mgr_dept_hours" in sheets_list:
        mgr_ids = _manager_emp_ids(full_emp_ids)
        mgr_dept_hours_rows = _build_manager_department_hours_rows(month, mgr_ids)
        headers = ["部门名称", "总工时（小时）", "部门人数"]
        row_data = []
        for r in mgr_dept_hours_rows:
            row_data.append([
                r.get("dept_name", ""),
                r.get("total_hours", 0),
                r.get("member_count", 0)
            ])
        headers, row_data = _filter_columns(headers, row_data, "mgr_dept_hours_headers")
        ws8 = wb.create_sheet("管理人员部门工时查询")
        ws8.append(headers)
        for row in row_data:
            ws8.append(row)

    if not wb.sheetnames:
        wb.create_sheet("无选择数据")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"汇总下载_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def manager_attendance_api():
    emp_ids = _accessible_manager_emp_ids()
    requested_ids = _requested_emp_ids()
    if requested_ids:
        allowed = set(emp_ids)
        emp_ids = [emp_id for emp_id in requested_ids if emp_id in allowed]
    options = _manager_options()
    rows = build_manager_rows(options, emp_ids)

    include_actual_attendance_days = request.args.get("show_actual_attendance_days") == "1"
    include_emp_no = request.args.get("show_emp_no") == "1"

    return jsonify(
        {
            "headers": manager_headers(include_actual_attendance_days, include_emp_no),
            "rows": rows_as_table(rows, include_actual_attendance_days, include_emp_no),
            "month": options.month,
            "factory_rest_days": options.factory_rest_days,
            "monthly_benefit_days": options.monthly_benefit_days,
        }
    )


def manager_overtime_query_api():
    from routes.admin_core import _manager_month_rows, _manager_overtime_values

    year = request.args.get("year", type=int) or datetime.now().year
    values = _manager_overtime_values(year)
    allowed_ids = set(_accessible_manager_emp_ids())
    values = {name: row for name, row in values.items() if row.get("emp_id") in allowed_ids}
    requested_ids = set(_requested_emp_ids())
    if requested_ids:
        values = {name: row for name, row in values.items() if row.get("emp_id") in requested_ids}
    return jsonify(
        {
            "year": year,
            "headers": ["部门", "姓名", "前年累积天数", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余调休天数", "备注"],
            "rows": _manager_month_rows(values, "剩余调休天数"),
        }
    )


def manager_annual_leave_query_api():
    from routes.admin_core import _annual_leave_value_keys, _manager_annual_leave_values, _manager_month_rows

    year = request.args.get("year", type=int) or datetime.now().year
    values = _manager_annual_leave_values(year)
    allowed_ids = set(_accessible_manager_emp_ids())
    values = {name: row for name, row in values.items() if row.get("emp_id") in allowed_ids}
    requested_ids = set(_requested_emp_ids())
    if requested_ids:
        values = {name: row for name, row in values.items() if row.get("emp_id") in requested_ids}
    return jsonify(
        {
            "year": year,
            "headers": ["部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"],
            "rows": _manager_month_rows(values, "剩余年休天数", _annual_leave_value_keys()),
        }
    )


def manager_department_hours_api():
    emp_ids = _accessible_manager_emp_ids()
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    return jsonify(_build_manager_department_hours_rows(month, emp_ids))


def manager_department_hours_export_api():
    emp_ids = _accessible_manager_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No manager assigned"}), 400

    month = _resolve_query_month()
    rows = _build_manager_department_hours_rows(month, emp_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "管理人员部门工时查询"
    ws.append(["部门名称", "总工时（小时）", "部门人数"])
    for row in rows:
        ws.append([row.get("dept_name", ""), row.get("total_hours", 0), row.get("member_count", 0)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员部门工时查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _copy_manager_template_row_style(ws, source_row_idx: int, target_row_idx: int) -> None:
    if source_row_idx == target_row_idx:
        return
    for col_idx in range(1, ws.max_column + 1):
        source = ws.cell(source_row_idx, col_idx)
        target = ws.cell(target_row_idx, col_idx)
        target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
    ws.row_dimensions[target_row_idx].height = ws.row_dimensions[source_row_idx].height


def _manager_export_rows_with_top_level_departments(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    normalized_rows: list[dict[str, object]] = []
    for item in rows:
        next_item = dict(item)
        emp_id = item.get("emp_id")
        employee = db.session.get(Employee, emp_id) if emp_id else None
        next_item["dept_name"] = _top_level_department_name(employee.department) if employee and employee.department else ""
        normalized_rows.append(next_item)
    return normalized_rows


def _manager_template_month_text(month: str) -> str:
    year, month_no = month.split("-", 1)
    return f"{int(year)}年{int(month_no)}月"


def _manager_template_title(month: str) -> str:
    return f"{_manager_template_month_text(month)}份考勤记录"


def _manager_template_current_month_text() -> str:
    return f"{datetime.now().year}年{datetime.now().month}月"


def _extract_template_header_footer_xml(template_path: Path) -> str | None:
    try:
        with ZipFile(template_path) as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    except (KeyError, OSError):
        return None

    start = sheet_xml.find("<headerFooter")
    if start < 0:
        return None
    end = sheet_xml.find("</headerFooter>", start)
    if end < 0:
        return None
    return sheet_xml[start : end + len("</headerFooter>")]


def _apply_sheet_header_footer_xml(workbook_bytes: bytes, header_footer_xml: str | None) -> bytes:
    if not header_footer_xml:
        return workbook_bytes

    source = BytesIO(workbook_bytes)
    target = BytesIO()
    with ZipFile(source) as src_zip, ZipFile(target, "w", compression=ZIP_DEFLATED) as dst_zip:
        for info in src_zip.infolist():
            data = src_zip.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                sheet_xml = data.decode("utf-8")
                start = sheet_xml.find("<headerFooter")
                if start >= 0:
                    end = sheet_xml.find("</headerFooter>", start)
                    if end >= 0:
                        sheet_xml = (
                            sheet_xml[:start]
                            + header_footer_xml
                            + sheet_xml[end + len("</headerFooter>") :]
                        )
                else:
                    sheet_xml = sheet_xml.replace("</worksheet>", f"{header_footer_xml}</worksheet>")
                data = sheet_xml.encode("utf-8")
            dst_zip.writestr(info, data)
    return target.getvalue()


def _apply_manager_department_header_style(ws, row_idx: int, template_row_idx: int = MANAGER_TEMPLATE_FIRST_DATA_ROW) -> None:
    source = ws.cell(template_row_idx, 1)
    target = ws.cell(row_idx, 1)
    target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format


def _clear_manager_template_merges(
    ws,
    start_row: int = MANAGER_TEMPLATE_FIRST_DATA_ROW,
    end_row: int = MANAGER_TEMPLATE_LAST_DATA_ROW,
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_col == 1
            and merged_range.max_col == 1
            and merged_range.max_row >= start_row
            and merged_range.min_row <= end_row
        ):
            ws.unmerge_cells(str(merged_range))


def _manager_template_values(item: dict[str, object], include_actual_attendance_days: bool) -> list[object]:
    return [
        item.get("name", ""),
        item.get("attendance_days", 0),
        *([item.get("actual_attendance_days", 0)] if include_actual_attendance_days else []),
        item.get("personal_sick_days", 0),
        item.get("injury_days", 0),
        item.get("business_trip_days", 0),
        item.get("marriage_days", 0),
        item.get("funeral_days", 0),
        item.get("late_early_minutes", 0),
        item.get("summary", ""),
        item.get("benefit_days", 0),
        item.get("overtime_change", 0),
        item.get("remark", ""),
    ]


def _fill_manager_template(ws, rows: list[dict[str, object]], month: str, include_actual_attendance_days: bool = True) -> None:
    headers = manager_headers(include_actual_attendance_days)
    template_data_rows = MANAGER_TEMPLATE_NOTICE_ROW - MANAGER_TEMPLATE_FIRST_DATA_ROW
    extra_rows = max(len(rows) - template_data_rows, 0)
    footer_start_row = MANAGER_TEMPLATE_NOTICE_ROW
    date_row = MANAGER_TEMPLATE_DATE_ROW + extra_rows

    if (
        include_actual_attendance_days
        and ws.max_column == len(MANAGER_HEADERS) - 1
        and ws.cell(MANAGER_TEMPLATE_HEADER_ROW, 4).value == "事/病假"
    ):
        ws.insert_cols(4)
        for row_idx in range(1, ws.max_row + 1):
            source = ws.cell(row_idx, 3)
            target = ws.cell(row_idx, 4)
            target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
    funeral_col_idx = 9 if include_actual_attendance_days else 8
    late_col_idx = funeral_col_idx + 1
    if ws.max_column == len(headers) - 1 and ws.cell(MANAGER_TEMPLATE_HEADER_ROW, funeral_col_idx).value == "迟到\\早退":
        ws.insert_cols(funeral_col_idx)
        for row_idx in range(1, ws.max_row + 1):
            source = ws.cell(row_idx, late_col_idx)
            target = ws.cell(row_idx, funeral_col_idx)
            target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format

    ws.cell(MANAGER_TEMPLATE_TITLE_ROW, 1).value = _manager_template_title(month)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(MANAGER_TEMPLATE_HEADER_ROW, col_idx).value = header

    first_data_row = MANAGER_TEMPLATE_FIRST_DATA_ROW
    _clear_manager_template_merges(ws, first_data_row, MANAGER_TEMPLATE_LAST_DATA_ROW)
    template_style_row = MANAGER_TEMPLATE_LAST_DATA_ROW
    current_data_rows = template_data_rows
    required_data_rows = len(rows)

    if required_data_rows > current_data_rows:
        extra_rows = required_data_rows - current_data_rows
        ws.insert_rows(footer_start_row, extra_rows)
        for target_row_idx in range(footer_start_row, footer_start_row + extra_rows):
            _copy_manager_template_row_style(ws, template_style_row, target_row_idx)
    elif required_data_rows < current_data_rows:
        for row_idx in range(first_data_row + required_data_rows, footer_start_row):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row_idx, col_idx).value = None

    clear_until_row = first_data_row + max(required_data_rows, current_data_rows) - 1
    for row_idx in range(first_data_row, clear_until_row + 1):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row_idx, col_idx).value = None

    previous_dept_name = None
    group_start_row = None
    merge_ranges: list[tuple[int, int]] = []

    for offset, item in enumerate(rows, start=0):
        row_idx = first_data_row + offset
        dept_name = str(item.get("dept_name", "") or "").strip()

        if dept_name != previous_dept_name:
            if group_start_row is not None and row_idx - 1 > group_start_row:
                merge_ranges.append((group_start_row, row_idx - 1))
            group_start_row = row_idx
            previous_dept_name = dept_name
            _apply_manager_department_header_style(ws, row_idx)
            ws.cell(row_idx, 1).value = dept_name

        values = _manager_template_values(item, include_actual_attendance_days)
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx).value = value

    final_data_row = first_data_row + required_data_rows - 1
    if group_start_row is not None and final_data_row > group_start_row:
        merge_ranges.append((group_start_row, final_data_row))

    for start_row, end_row in merge_ranges:
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            start_cell = ws.cell(start_row, 1)
            left_style = copy(start_cell.border.left)
            right_style = copy(start_cell.border.right)
            top_style = copy(start_cell.border.top)
            bottom_style = copy(start_cell.border.bottom)
            middle_top_style = copy(ws.cell(start_row + 1, 1).border.top)
            middle_bottom_style = copy(ws.cell(end_row - 1, 1).border.bottom)
            for row_idx in range(start_row, end_row + 1):
                cell = ws.cell(row_idx, 1)
                cell.alignment = copy(start_cell.alignment)
                cell.border = openpyxl.styles.Border(
                    left=left_style,
                    right=right_style,
                    top=top_style if row_idx == start_row else middle_top_style,
                    bottom=bottom_style if row_idx == end_row else middle_bottom_style,
                )

    ws.cell(date_row, 10).value = _manager_template_current_month_text()
    ws.print_area = f"A1:M{ws.max_row}"


def manager_attendance_export_api():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    requested_ids = _requested_emp_ids()
    if requested_ids:
        allowed = set(emp_ids)
        emp_ids = [emp_id for emp_id in requested_ids if emp_id in allowed]
    options = _manager_options()
    rows = _manager_export_rows_with_top_level_departments(build_manager_rows(options, emp_ids))
    include_actual_attendance_days = request.args.get("show_actual_attendance_days") == "1"
    include_emp_no = request.args.get("show_emp_no") == "1"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "管理人员查询"
    ws.append(manager_headers(include_actual_attendance_days, include_emp_no))
    for row in rows_as_table(rows, include_actual_attendance_days, include_emp_no):
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员考勤查询_{options.month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def manager_attendance_template_export_api():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    requested_ids = _requested_emp_ids()
    if requested_ids:
        allowed = set(emp_ids)
        emp_ids = [emp_id for emp_id in requested_ids if emp_id in allowed]
    options = _manager_options()
    rows = _manager_export_rows_with_top_level_departments(build_manager_rows(options, emp_ids))
    include_actual_attendance_days = request.args.get("show_actual_attendance_days") == "1"

    if MANAGER_ATTENDANCE_TEMPLATE_PATH.exists():
        wb = openpyxl.load_workbook(MANAGER_ATTENDANCE_TEMPLATE_PATH)
        ws = wb.active
        _fill_manager_template(ws, rows, options.month, include_actual_attendance_days)
        header_footer_xml = _extract_template_header_footer_xml(MANAGER_ATTENDANCE_TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "管理人员查询"
        ws.append(manager_headers(include_actual_attendance_days))
        for row in rows_as_table(rows, include_actual_attendance_days):
            ws.append(row)
        header_footer_xml = None

    output = BytesIO()
    wb.save(output)
    output_bytes = _apply_sheet_header_footer_xml(output.getvalue(), header_footer_xml)
    return send_file(
        BytesIO(output_bytes),
        as_attachment=True,
        download_name=f"管理人员考勤查询_{options.month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def abnormal_attendance_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    return jsonify(_build_abnormal_rows(month, emp_ids))


def abnormal_attendance_export_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    rows = _build_abnormal_rows(month, emp_ids)

    headers = ["部门名称", "人员编号", "人员姓名", "异常考勤次数"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工异常查询"
    ws.append(headers)
    for r in rows:
        ws.append([r.get("dept_name", ""), r.get("emp_no", ""), r.get("name", ""), r.get("abnormal_count", 0)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"员工异常查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def department_hours_api():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    return jsonify(_build_department_hours_rows(month, emp_ids))


def department_hours_export_api():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    rows = _build_department_hours_rows(month, emp_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工部门工时查询"
    ws.append(["部门名称", "总工时（小时）", "部门人数"])
    for row in rows:
        ws.append([row.get("dept_name", ""), row.get("total_hours", 0), row.get("member_count", 0)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"员工部门工时查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
