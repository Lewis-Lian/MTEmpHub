import io
import os
from datetime import date, timedelta
from pathlib import Path
from types import SimpleNamespace
from zipfile import ZipFile
import tempfile
import unittest
import warnings

import openpyxl
from flask import Flask, g

from models import db
from models.account_set import AccountSet, AccountSetFactoryRestDay
from models.daily_record import DailyRecord
from models.department import Department
from models.employee import Employee
from models.manager_month_stat import ManagerMonthStat
from models.user import User
from routes import register_routes
from routes.admin import _factory_rest_unit, _manager_attendance_options
from routes.employee import _fill_manager_template, _manager_options
from services.bootstrap_service import ensure_schema_compatibility
from utils.app_navigation import nav_context, visible_modules


class AttendanceOverrideFeatureTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmpdir.name, "test.db")
        self.upload_dir = os.path.join(self.tmpdir.name, "uploads")
        project_root = os.path.dirname(os.path.dirname(__file__))

        self.app = Flask(
            __name__,
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        self.app.config.update(
            TESTING=True,
            SECRET_KEY="test-secret",
            SQLALCHEMY_DATABASE_URI=f"sqlite:///{self.db_path}",
            SQLALCHEMY_TRACK_MODIFICATIONS=False,
            JWT_EXPIRES_DELTA=timedelta(hours=12),
            FRONTEND_ORIGIN="http://localhost:5173",
            UPLOAD_FOLDER=self.upload_dir,
        )
        os.makedirs(self.upload_dir, exist_ok=True)

        db.init_app(self.app)
        register_routes(self.app)

        with self.app.app_context():
            db.create_all()
            admin = User(username="admin", role="admin")
            admin.set_password("admin123")
            dept = Department(dept_no="D001", dept_name="行政部")
            db.session.add_all([admin, dept])
            db.session.flush()

            employee = Employee(emp_no="E001", name="员工甲", dept_id=dept.id, is_manager=False)
            employee_b = Employee(emp_no="E002", name="员工乙", dept_id=dept.id, is_manager=False)
            manager = Employee(emp_no="M001", name="经理甲", dept_id=dept.id, is_manager=True)
            db.session.add_all([employee, employee_b, manager])
            db.session.add(AccountSet(month="2026-05", name="2026-05", is_active=True, is_locked=False))
            db.session.commit()

            self.employee_id = employee.id
            self.employee_b_id = employee_b.id
            self.manager_id = manager.id

        self.client = self.app.test_client()
        self.client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})

    def tearDown(self) -> None:
        with self.app.app_context():
            db.session.remove()
            db.drop_all()
        self.tmpdir.cleanup()

    def _manager_template_path(self) -> Path:
        project_root = Path(__file__).resolve().parents[1]
        return project_root / "templates" / "export_templates" / "manager_attendance.xlsx"

    def test_employee_manual_save_creates_history_once(self) -> None:
        payload = {
            "month": "2026-05",
            "emp_id": self.employee_id,
            "attendance_days": "3",
            "work_hours": "21.5",
            "half_days": "1",
            "late_early_minutes": "10",
            "remark": "手工修正",
        }

        first = self.client.put("/api/admin/employee-attendance-overrides/record", json=payload)
        self.assertEqual(first.status_code, 200)

        history = self.client.get(
            f"/api/admin/employee-attendance-overrides/history?emp_id={self.employee_id}&month=2026-05"
        )
        self.assertEqual(history.status_code, 200)
        first_rows = history.get_json()["rows"]
        self.assertEqual(len(first_rows), 1)
        self.assertEqual(first_rows[0]["action_type"], "manual_save")

        second = self.client.put("/api/admin/employee-attendance-overrides/record", json=payload)
        self.assertEqual(second.status_code, 200)
        history_again = self.client.get(
            f"/api/admin/employee-attendance-overrides/history?emp_id={self.employee_id}&month=2026-05"
        )
        self.assertEqual(len(history_again.get_json()["rows"]), 1)

    def test_employee_override_list_returns_selected_employees(self) -> None:
        self.client.put(
            "/api/admin/employee-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.employee_id,
                "attendance_days": "2",
                "work_hours": "15.5",
                "remark": "员工甲修正",
            },
        )
        self.client.put(
            "/api/admin/employee-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.employee_b_id,
                "attendance_days": "4",
                "half_days": "1",
                "remark": "员工乙修正",
            },
        )

        res = self.client.get(
            f"/api/admin/employee-attendance-overrides?month=2026-05&emp_ids={self.employee_id},{self.employee_b_id}"
        )
        self.assertEqual(res.status_code, 200)
        rows = res.get_json()["rows"]
        self.assertEqual(len(rows), 2)
        self.assertEqual([row["employee"]["emp_no"] for row in rows], ["E001", "E002"])
        self.assertEqual(rows[0]["override"]["attendance_days"], 2.0)
        self.assertEqual(rows[0]["override"]["work_hours"], 15.5)
        self.assertEqual(rows[1]["override"]["attendance_days"], 4.0)
        self.assertEqual(rows[1]["override"]["half_days"], 1)

    def test_manager_override_list_returns_selected_employees(self) -> None:
        self.client.put(
            "/api/admin/manager-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.manager_id,
                "attendance_days": "20",
                "injury_days": "1",
                "remark": "经理修正",
            },
        )

        res = self.client.get(f"/api/admin/manager-attendance-overrides?month=2026-05&emp_ids={self.manager_id}")
        self.assertEqual(res.status_code, 200)
        rows = res.get_json()["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["employee"]["emp_no"], "M001")
        self.assertEqual(rows[0]["override"]["attendance_days"], 20.0)
        self.assertEqual(rows[0]["override"]["injury_days"], 1.0)
        self.assertEqual(rows[0]["override"]["remark"], "经理修正")

    def test_employee_dashboard_api_uses_override_attendance_days(self) -> None:
        with self.app.app_context():
            db.session.add(
                DailyRecord(
                    emp_id=self.employee_id,
                    record_date=date(2026, 5, 6),
                    expected_hours=8,
                    actual_hours=8,
                    absent_hours=0,
                    leave_hours=0,
                    overtime_hours=0,
                    late_minutes=0,
                    early_leave_minutes=0,
                    employee_payload={
                        "expected_hours": 8,
                        "actual_hours": 8,
                        "absent_hours": 0,
                        "leave_hours": 0,
                        "overtime_hours": 0,
                        "late_minutes": 0,
                        "early_leave_minutes": 0,
                    },
                )
            )
            db.session.commit()

        self.client.put(
            "/api/admin/employee-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.employee_id,
                "attendance_days": "3",
                "remark": "查询修正",
            },
        )

        res = self.client.get(f"/api/query/employee-dashboard?month=2026-05&emp_ids={self.employee_id}")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0][3], 3.0)

    def test_punch_records_export_filters_to_requested_headers(self) -> None:
        with self.app.app_context():
            db.session.add(
                DailyRecord(
                    emp_id=self.employee_id,
                    record_date=date(2026, 5, 6),
                    actual_hours=8,
                    check_in_times=["08:00"],
                    check_out_times=["17:00"],
                    late_minutes=15,
                    early_leave_minutes=5,
                    exception_reason="迟到",
                    raw_data={"打卡记录": ["08:00", "17:00"]},
                    employee_payload={
                        "actual_hours": 8,
                        "check_in_times": ["08:00"],
                        "check_out_times": ["17:00"],
                        "late_minutes": 15,
                        "early_leave_minutes": 5,
                        "exception_reason": "迟到",
                        "raw_data": {"打卡记录": ["08:00", "17:00"]},
                    },
                )
            )
            db.session.commit()

        res = self.client.get(
            f"/api/query/punch-records/export?month=2026-05&emp_ids={self.employee_id}"
            "&punch_headers=日期,员工编号,原始打卡数据,打卡次数"
        )
        self.assertEqual(res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        ws = wb.active
        self.assertEqual(ws.title, "打卡数据查询")
        self.assertEqual(
            [ws.cell(1, col_idx).value for col_idx in range(1, 5)],
            ["日期", "员工编号", "原始打卡数据", "打卡次数"],
        )
        self.assertEqual(ws.max_column, 4)
        self.assertEqual(ws["A2"].value, "2026-05-06")
        self.assertEqual(ws["B2"].value, "E001")
        self.assertIn("08:00", ws["C2"].value)
        self.assertEqual(ws["D2"].value, 2)

    def test_query_bootstrap_marks_legacy_factory_rest_days_as_requires_detail(self) -> None:
        with self.app.app_context():
            account_set = AccountSet.query.filter_by(month="2026-05").first()
            self.assertIsNotNone(account_set)
            account_set.factory_rest_days = 2.0
            db.session.commit()

        res = self.client.get("/api/query/bootstrap")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        account_set_payload = payload["account_sets"][0]
        self.assertEqual(account_set_payload["factory_rest_days"], 0)
        self.assertTrue(account_set_payload["factory_rest_requires_detail"])
        self.assertEqual(account_set_payload["legacy_factory_rest_days"], 2.0)

    def test_home_summary_api_returns_empty_when_no_account_set(self) -> None:
        with self.app.app_context():
            AccountSet.query.delete()
            db.session.commit()

        viewer_client = self.app.test_client()
        with self.app.app_context():
            viewer = User(username="viewer", role="readonly", page_permissions={"query_home": True})
            viewer.set_password("viewer123")
            db.session.add(viewer)
            db.session.commit()
        viewer_client.post("/api/auth/login", json={"username": "viewer", "password": "viewer123"})
        res = viewer_client.get("/api/query/home-summary")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertFalse(payload["has_data"])
        self.assertEqual(payload["empty_state"], "暂无账套，暂无数据")

    def test_home_summary_api_returns_manager_summary_by_profile_emp_no(self) -> None:
        with self.app.app_context():
            viewer = User(username="viewer", role="readonly", page_permissions={"query_home": True})
            viewer.profile_emp_no = "M001"
            viewer.profile_name = "经理甲账号"
            viewer.set_password("viewer123")
            db.session.add(viewer)
            db.session.add(AccountSet(month="2026-04", name="2026-04", is_active=False, is_locked=False))
            db.session.add(AccountSet(month="2026-06", name="2026-06", is_active=False, is_locked=False))
            db.session.add(ManagerMonthStat(emp_id=self.manager_id, year=2026, stat_type="annual_leave", m4=1, m5=2))
            db.session.add(ManagerMonthStat(emp_id=self.manager_id, year=2026, stat_type="overtime", prev_dec=2, m4=1, m5=-0.5, m6=3))
            db.session.commit()

        viewer_client = self.app.test_client()
        viewer_client.post("/api/auth/login", json={"username": "viewer", "password": "viewer123"})
        april_payload = viewer_client.get("/api/query/home-summary?month=2026-04").get_json()
        self.assertTrue(april_payload["has_data"])
        self.assertEqual(april_payload["month"], "2026-04")
        self.assertEqual(april_payload["manager"]["emp_no"], "M001")
        self.assertEqual(april_payload["manager"]["name"], "经理甲")
        self.assertEqual(april_payload["manager"]["dept_name"], "行政部")
        self.assertEqual(april_payload["summary"]["benefit_days"], 11)
        self.assertEqual(april_payload["summary"]["overtime_remaining_days"], 3)

        may_payload = viewer_client.get("/api/query/home-summary?month=2026-05").get_json()
        self.assertEqual(may_payload["summary"]["benefit_days"], 9)
        self.assertEqual(may_payload["summary"]["overtime_remaining_days"], 2.5)

    def test_manager_attendance_template_is_stored_in_repo(self) -> None:
        template_path = self._manager_template_path()
        self.assertTrue(template_path.exists())

        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb = openpyxl.load_workbook(template_path)
        self.assertEqual([str(item.message) for item in caught], [])
        ws = wb.active
        self.assertEqual(ws.print_title_rows, "$1:$2")
        self.assertTrue(ws.print_options.horizontalCentered)

        with ZipFile(template_path) as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
        self.assertIn("<oddFooter>", sheet_xml)
        self.assertIn("第 &amp;P 页，共 &amp;N 页", sheet_xml)

    def test_fill_manager_template_clears_unused_sample_rows(self) -> None:
        wb = openpyxl.load_workbook(self._manager_template_path())
        ws = wb.active
        current_month_text = __import__("datetime").datetime.now().strftime("%Y年%-m月")

        _fill_manager_template(
            ws,
            [
                {
                    "dept_name": "行政部",
                    "name": "经理甲",
                    "attendance_days": 20,
                    "personal_sick_days": 0,
                    "injury_days": 0,
                    "business_trip_days": 0,
                    "marriage_days": 0,
                    "funeral_days": 0,
                    "late_early_minutes": 0,
                    "summary": "",
                    "benefit_days": 0,
                    "overtime_change": 0,
                    "remark": "",
                }
            ],
            "2026-04",
            include_actual_attendance_days=False,
        )

        self.assertEqual(ws["A1"].value, "2026年4月份考勤记录")
        self.assertEqual(ws["J109"].value, current_month_text)
        self.assertEqual(ws["A3"].value, "行政部")
        self.assertEqual(ws["B3"].value, "经理甲")
        self.assertIsNone(ws["B4"].value)

    def test_manager_attendance_options_use_zero_for_legacy_factory_rest_days_without_detail(self) -> None:
        with self.app.app_context():
            account_set = AccountSet.query.filter_by(month="2026-05").first()
            self.assertIsNotNone(account_set)
            account_set.factory_rest_days = 2.0
            db.session.commit()

        with self.app.test_request_context("/employee/api/manager-attendance?month=2026-05"):
            options = _manager_options()
            self.assertEqual(options.factory_rest_days, 0)

        with self.app.app_context():
            admin_options = _manager_attendance_options("2026-05")
            self.assertEqual(admin_options.factory_rest_days, 0)

    def test_factory_rest_unit_rejects_unknown_period(self) -> None:
        with self.assertRaises(ValueError):
            _factory_rest_unit("invalid")

    def test_ensure_schema_compatibility_is_idempotent_for_factory_rest_table(self) -> None:
        with self.app.app_context():
            ensure_schema_compatibility()
            ensure_schema_compatibility()

            account_set = AccountSet.query.filter_by(month="2026-05").first()
            self.assertIsNotNone(account_set)
            db.session.add(
                AccountSetFactoryRestDay(
                    account_set_id=account_set.id,
                    rest_date=date(2026, 5, 8),
                    rest_period="full",
                )
            )
            db.session.commit()

            refreshed = db.session.get(AccountSet, account_set.id)
            self.assertEqual(len(refreshed.factory_rest_entries), 1)

    def test_product_navigation_filters_readonly_permissions(self) -> None:
        readonly_user = SimpleNamespace(
            username="reader",
            role="readonly",
            has_any_page_access=lambda keys: "employee_dashboard" in keys,
            can_access_page=lambda key: key in ("employee_dashboard", "query_home"),
        )

        modules = visible_modules(readonly_user)
        query_module = next(module for module in modules if module["slug"] == "query")
        entry_keys = [entry["key"] for entry in query_module["entries"]]
        self.assertIn("employee_dashboard", entry_keys)
        self.assertNotIn("manager_query", entry_keys)

        context = nav_context(readonly_user, "/employee/home")
        self.assertEqual([entry["href"] for entry in context["current_entries"]], ["/employee/dashboard"])


if __name__ == "__main__":
    unittest.main()
