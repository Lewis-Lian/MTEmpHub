import os
import tempfile
import unittest
from datetime import datetime, timedelta
from io import BytesIO

from flask import Flask
import openpyxl

from models import db
from models.account_set import AccountSet
from models.department import Department
from models.employee import Employee
from models.employee_attendance_override import EmployeeAttendanceOverride
from models.manager_attendance_override import ManagerAttendanceOverride
from models.shift import Shift
from models.user import User
from routes import register_routes


class ApiAdminTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)
        self.app = Flask(__name__)
        self.app.config.update(
            TESTING=True,
            SECRET_KEY="test-secret",
            SQLALCHEMY_DATABASE_URI=f"sqlite:///{self.tmpdir.name}/api-admin.db",
            SQLALCHEMY_TRACK_MODIFICATIONS=False,
            JWT_EXPIRES_DELTA=timedelta(hours=12),
            FRONTEND_ORIGIN="http://localhost:5173",
            SESSION_COOKIE_NAME="api_admin_access_token",
            SESSION_COOKIE_SAMESITE="None",
            SESSION_COOKIE_SECURE=False,
            UPLOAD_FOLDER=os.path.join(self.tmpdir.name, "uploads"),
        )
        os.makedirs(self.app.config["UPLOAD_FOLDER"], exist_ok=True)
        db.init_app(self.app)
        register_routes(self.app)

        with self.app.app_context():
            db.create_all()

            admin = User(username="admin", role="admin")
            admin.set_password("admin123")
            dept = Department(dept_no="D001", dept_name="行政部")
            alt_dept = Department(dept_no="D002", dept_name="生产部")
            shift = Shift(
                shift_no="S001",
                shift_name="白班",
                time_slots=[{"start": "08:00", "end": "17:00"}],
                is_cross_day=False,
            )
            db.session.add_all([admin, dept, alt_dept, shift])
            db.session.flush()

            employee = Employee(emp_no="E001", name="员工甲", dept_id=dept.id, is_manager=False)
            manager = Employee(emp_no="M001", name="经理甲", dept_id=dept.id, is_manager=True)
            nursing_employee = Employee(
                emp_no="E002",
                name="筛选目标",
                dept_id=alt_dept.id,
                is_manager=False,
                is_nursing=True,
                employee_stats_attendance_source="auto_fallback",
                manager_stats_attendance_source="employee",
            )
            db.session.add_all([employee, manager, nursing_employee])
            db.session.add(AccountSet(month="2026-05", name="2026-05", is_active=True, is_locked=False))
            db.session.commit()

            self.employee_id = employee.id
            self.manager_id = manager.id
            self.filtered_employee_id = nursing_employee.id

        self.client = self.app.test_client()

    def tearDown(self) -> None:
        with self.app.app_context():
            db.session.remove()
            db.drop_all()

    def _login(self) -> None:
        self.client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})

    def _xlsx_file(self, rows: list[list[object]], filename: str) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        output.name = filename
        return output

    def test_admin_routes_are_registered_under_api_prefix(self) -> None:
        rules = {rule.rule for rule in self.app.url_map.iter_rules()}

        self.assertTrue(
            {
                "/api/admin/bootstrap",
                "/api/admin/accounts",
                "/api/admin/employees",
                "/api/admin/departments",
                "/api/admin/shifts",
                "/api/admin/employee-attendance-overrides",
                "/api/admin/employee-attendance-overrides/history",
                "/api/admin/employee-attendance-overrides/record",
                "/api/admin/employee-attendance-overrides/template",
                "/api/admin/employee-attendance-overrides/export",
                "/api/admin/employee-attendance-overrides/import",
                "/api/admin/manager-attendance-overrides",
                "/api/admin/manager-attendance-overrides/history",
                "/api/admin/manager-attendance-overrides/record",
                "/api/admin/manager-attendance-overrides/template",
                "/api/admin/manager-attendance-overrides/export",
                "/api/admin/manager-attendance-overrides/import",
                "/api/admin/manager-overtime",
                "/api/admin/manager-overtime/records",
                "/api/admin/manager-overtime/template",
                "/api/admin/manager-overtime/export",
                "/api/admin/manager-overtime/import",
                "/api/admin/manager-annual-leave",
                "/api/admin/manager-annual-leave/records",
                "/api/admin/manager-annual-leave/template",
                "/api/admin/manager-annual-leave/export",
                "/api/admin/manager-annual-leave/import",
                "/api/admin/disabled-users",
                "/api/admin/database-settings",
                "/api/admin/disabled-users/<int:user_id>/unlock",
                "/api/admin/users/<int:user_id>",
            }.issubset(rules)
        )

    def test_admin_bootstrap_requires_login(self) -> None:
        response = self.client.get("/api/admin/bootstrap")

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.get_json(), {"error": "Unauthorized"})

    def test_admin_bootstrap_returns_departments_and_shifts(self) -> None:
        self._login()

        response = self.client.get("/api/admin/bootstrap")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(sorted(row["dept_no"] for row in payload["departments"]), ["D001", "D002"])
        self.assertEqual([row["shift_no"] for row in payload["shifts"]], ["S001"])

    def test_database_settings_returns_masked_database_summary(self) -> None:
        self._login()

        response = self.client.get("/api/admin/database-settings")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertIsInstance(payload, dict)
        self.assertIn("current", payload)
        labels = {row["item"]: row["value"] for row in payload["current"]}
        self.assertEqual(labels["数据库类型"], "sqlite")
        self.assertEqual(labels["用户名"], "-")
        self.assertIn("api-admin.db", labels["数据库名称"])

    def test_admin_collection_endpoints_return_expected_payloads(self) -> None:
        self._login()

        accounts_response = self.client.get("/api/admin/accounts")
        employees_response = self.client.get("/api/admin/employees")
        departments_response = self.client.get("/api/admin/departments")
        shifts_response = self.client.get("/api/admin/shifts")
        employee_override_response = self.client.get(
            f"/api/admin/employee-attendance-overrides?month=2026-05&emp_ids={self.employee_id}"
        )
        employee_history_response = self.client.get("/api/admin/employee-attendance-overrides/history?month=2026-05")
        manager_override_response = self.client.get(
            f"/api/admin/manager-attendance-overrides?month=2026-05&emp_ids={self.manager_id}"
        )
        manager_history_response = self.client.get("/api/admin/manager-attendance-overrides/history?month=2026-05")
        manager_overtime_response = self.client.get(
            f"/api/admin/manager-overtime?year=2026&emp_ids={self.manager_id}"
        )
        manager_annual_leave_response = self.client.get(
            f"/api/admin/manager-annual-leave?year=2026&emp_ids={self.manager_id}"
        )
        accounts_payload = accounts_response.get_json()
        employees_payload = employees_response.get_json()
        departments_payload = departments_response.get_json()
        shifts_payload = shifts_response.get_json()
        employee_override_payload = employee_override_response.get_json()
        employee_history_payload = employee_history_response.get_json()
        manager_override_payload = manager_override_response.get_json()
        manager_history_payload = manager_history_response.get_json()
        manager_overtime_payload = manager_overtime_response.get_json()
        manager_annual_leave_payload = manager_annual_leave_response.get_json()

        self.assertEqual(accounts_response.status_code, 200)
        self.assertIsInstance(accounts_payload, list)
        self.assertIn("username", accounts_payload[0])
        self.assertIn("role", accounts_payload[0])
        self.assertEqual(accounts_payload[0]["username"], "admin")
        self.assertEqual(employees_response.status_code, 200)
        self.assertIsInstance(employees_payload, list)
        self.assertIn("emp_no", employees_payload[0])
        self.assertIn("name", employees_payload[0])
        self.assertEqual(employees_payload[0]["emp_no"], "E001")
        self.assertEqual(departments_response.status_code, 200)
        self.assertIsInstance(departments_payload, list)
        self.assertIn("dept_no", departments_payload[0])
        self.assertIn("dept_name", departments_payload[0])
        self.assertIn("D001", [row["dept_no"] for row in departments_payload])
        self.assertEqual(shifts_response.status_code, 200)
        self.assertIsInstance(shifts_payload, list)
        self.assertIn("shift_no", shifts_payload[0])
        self.assertIn("shift_name", shifts_payload[0])
        self.assertEqual(shifts_payload[0]["shift_no"], "S001")
        self.assertEqual(employee_override_response.status_code, 200)
        self.assertIn("rows", employee_override_payload)
        self.assertIn("month", employee_override_payload)
        self.assertIn("employee", employee_override_payload["rows"][0])
        self.assertIn("override", employee_override_payload["rows"][0])
        self.assertIn("applied", employee_override_payload["rows"][0])
        self.assertEqual(employee_history_response.status_code, 200)
        self.assertIn("rows", employee_history_payload)
        self.assertEqual(employee_history_payload["rows"], [])
        self.assertEqual(manager_override_response.status_code, 200)
        self.assertIn("rows", manager_override_payload)
        self.assertIn("month", manager_override_payload)
        self.assertIn("employee", manager_override_payload["rows"][0])
        self.assertIn("override", manager_override_payload["rows"][0])
        self.assertIn("applied", manager_override_payload["rows"][0])
        self.assertEqual(manager_history_response.status_code, 200)
        self.assertIn("rows", manager_history_payload)
        self.assertEqual(manager_history_payload["rows"], [])
        self.assertEqual(manager_overtime_response.status_code, 200)
        self.assertIsInstance(manager_overtime_payload, list)
        self.assertIn("name", manager_overtime_payload[0])
        self.assertIn("remark", manager_overtime_payload[0])
        self.assertEqual(manager_annual_leave_response.status_code, 200)
        self.assertIsInstance(manager_annual_leave_payload, list)
        self.assertIn("name", manager_annual_leave_payload[0])
        self.assertIn("remark", manager_annual_leave_payload[0])

    def test_employees_export_supports_current_filters(self) -> None:
        self._login()

        response = self.client.get(
            "/api/admin/employees/export"
            "?keyword=%E7%AD%9B%E9%80%89"
            "&type=employee"
            "&is_nursing=1"
            "&employee_source=auto_fallback"
            "&manager_source=employee"
            f"&ids={self.filtered_employee_id},{self.employee_id}"
        )

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.data))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], (
            "人员编号",
            "人员姓名",
            "部门名称",
            "班次编号",
            "是否管理人员",
            "是否哺乳假",
            "员工考勤统计来源",
            "管理人员考勤统计来源",
        ))
        self.assertEqual(
            rows[1:],
            [("E002", "筛选目标", "生产部", None, "否", "是", "自动回退", "员工考勤源文件取值")],
        )

    def test_manager_month_stat_record_endpoints_match_collection_payloads(self) -> None:
        self._login()

        overtime_collection = self.client.get(f"/api/admin/manager-overtime?year=2026&emp_ids={self.manager_id}")
        overtime_records = self.client.get(f"/api/admin/manager-overtime/records?year=2026&emp_ids={self.manager_id}")
        annual_leave_collection = self.client.get(
            f"/api/admin/manager-annual-leave?year=2026&emp_ids={self.manager_id}"
        )
        annual_leave_records = self.client.get(
            f"/api/admin/manager-annual-leave/records?year=2026&emp_ids={self.manager_id}"
        )

        self.assertEqual(overtime_collection.status_code, 200)
        self.assertEqual(overtime_records.status_code, 200)
        self.assertEqual(overtime_records.get_json(), overtime_collection.get_json())
        self.assertEqual(annual_leave_collection.status_code, 200)
        self.assertEqual(annual_leave_records.status_code, 200)
        self.assertEqual(annual_leave_records.get_json(), annual_leave_collection.get_json())

    def test_admin_can_list_and_unlock_disabled_users(self) -> None:
        with self.app.app_context():
            temp_locked = User(
                username="temp-locked",
                role="readonly",
                login_failed_attempts=5,
                login_locked_until=datetime.utcnow() + timedelta(minutes=10),
            )
            temp_locked.set_password("temp123")
            disabled = User(
                username="disabled-user",
                role="readonly",
                login_failed_attempts=10,
                login_disabled_until_admin_unlock=True,
                login_disabled_reason="too_many_failed_attempts",
            )
            disabled.set_password("disabled123")
            db.session.add_all([temp_locked, disabled])
            db.session.commit()
            disabled_id = disabled.id

        self._login()

        list_response = self.client.get("/api/admin/disabled-users")
        self.assertEqual(list_response.status_code, 200)
        payload = list_response.get_json()
        self.assertEqual([row["username"] for row in payload], ["disabled-user", "temp-locked"])
        self.assertTrue(payload[0]["login_disabled_until_admin_unlock"])
        self.assertIsNotNone(payload[1]["login_locked_until"])

        unlock_response = self.client.post(f"/api/admin/disabled-users/{disabled_id}/unlock")
        self.assertEqual(unlock_response.status_code, 200)

        with self.app.app_context():
            unlocked = db.session.get(User, disabled_id)
            self.assertEqual(unlocked.login_failed_attempts, 0)
            self.assertIsNone(unlocked.login_locked_until)
            self.assertFalse(unlocked.login_disabled_until_admin_unlock)
            self.assertIsNone(unlocked.login_disabled_reason)

    def test_admin_can_update_user(self) -> None:
        with self.app.app_context():
            user = User(username="user-to-edit", role="readonly")
            user.set_password("pwd123")
            db.session.add(user)
            db.session.commit()
            user_id = user.id

        self._login()

        dept_id = self.client.get("/api/admin/departments").get_json()[0]["id"]

        response = self.client.put(
            f"/api/admin/users/{user_id}",
            json={
                "role": "admin",
                "profile_emp_no": "E999",
                "profile_name": "修改后的名字",
                "profile_dept_id": dept_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["status"], "ok")

        with self.app.app_context():
            updated = db.session.get(User, user_id)
            self.assertEqual(updated.role, "admin")
            self.assertEqual(updated.profile_emp_no, "E999")
            self.assertEqual(updated.profile_name, "修改后的名字")

    def test_admin_template_download_endpoints_return_excel_attachments(self) -> None:
        self._login()

        departments_response = self.client.get("/api/admin/departments/template")
        employees_response = self.client.get("/api/admin/employees/template")

        self.assertEqual(departments_response.status_code, 200)
        self.assertEqual(employees_response.status_code, 200)
        self.assertIn(
            "attachment; filename=",
            departments_response.headers.get("Content-Disposition", ""),
        )
        self.assertIn(
            "attachment; filename=",
            employees_response.headers.get("Content-Disposition", ""),
        )
        self.assertEqual(
            departments_response.headers.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(
            employees_response.headers.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_admin_record_wrappers_reuse_existing_write_behavior(self) -> None:
        self._login()

        employee_response = self.client.put(
            "/api/admin/employee-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.employee_id,
                "attendance_days": "3",
                "work_hours": "21.5",
                "half_days": "1",
                "late_early_minutes": "10",
                "remark": "员工修正",
            },
        )
        manager_response = self.client.put(
            "/api/admin/manager-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.manager_id,
                "attendance_days": "20",
                "injury_days": "1",
                "business_trip_days": "2",
                "marriage_days": "0",
                "funeral_days": "0",
                "late_early_minutes": "5",
                "remark": "经理修正",
            },
        )

        self.assertEqual(employee_response.status_code, 200)
        self.assertEqual(manager_response.status_code, 200)

        employee_record = self.client.get(
            f"/api/admin/employee-attendance-overrides?month=2026-05&emp_ids={self.employee_id}"
        ).get_json()["rows"][0]
        manager_record = self.client.get(
            f"/api/admin/manager-attendance-overrides?month=2026-05&emp_ids={self.manager_id}"
        ).get_json()["rows"][0]
        employee_history = self.client.get(
            "/api/admin/employee-attendance-overrides/history?month=2026-05"
        ).get_json()["rows"]
        manager_history = self.client.get(
            "/api/admin/manager-attendance-overrides/history?month=2026-05"
        ).get_json()["rows"]

        self.assertEqual(employee_record["override"]["attendance_days"], 3.0)
        self.assertEqual(employee_record["override"]["remark"], "员工修正")
        self.assertEqual(manager_record["override"]["attendance_days"], 20.0)
        self.assertEqual(manager_record["override"]["remark"], "经理修正")
        self.assertEqual(employee_history[0]["action_type"], "manual_save")
        self.assertEqual(manager_history[0]["action_type"], "manual_save")

    def test_admin_import_wrappers_apply_department_and_employee_xlsx_files(self) -> None:
        self._login()

        department_file = self._xlsx_file(
            [
                ["部门编号", "部门名称", "上级部门编号"],
                ["D010", "生产中心", ""],
                ["D011", "生产一部", "D010"],
            ],
            "departments.xlsx",
        )
        department_response = self.client.post(
            "/api/admin/departments/import",
            data={"file": (department_file, "departments.xlsx")},
            content_type="multipart/form-data",
        )

        employee_file = self._xlsx_file(
            [
                [
                    "人员编号",
                    "人员姓名",
                    "部门名称",
                    "班次编号",
                    "是否管理人员",
                    "是否哺乳假",
                    "员工考勤统计来源",
                    "管理人员考勤统计来源",
                ],
                ["E010", "员工乙", "生产一部", "S001", "否", "是", "自动回退", "员工考勤源文件取值"],
            ],
            "employees.xlsx",
        )
        employee_response = self.client.post(
            "/api/admin/employees/import",
            data={"file": (employee_file, "employees.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(department_response.status_code, 200)
        self.assertEqual(department_response.get_json()["imported"], 2)
        self.assertEqual(employee_response.status_code, 200)
        self.assertEqual(employee_response.get_json()["imported"], 1)
        with self.app.app_context():
            parent = Department.query.filter_by(dept_no="D010").first()
            child = Department.query.filter_by(dept_no="D011").first()
            employee = Employee.query.filter_by(emp_no="E010").first()

            self.assertIsNotNone(parent)
            self.assertIsNotNone(child)
            self.assertEqual(child.parent_id, parent.id)
            self.assertIsNotNone(employee)
            self.assertEqual(employee.department.dept_name, "生产一部")
            self.assertEqual(employee.shift_assignment.shift.shift_no, "S001")
            self.assertTrue(employee.is_nursing)
            self.assertEqual(employee.employee_stats_attendance_source, "auto_fallback")
            self.assertEqual(employee.manager_stats_attendance_source, "employee")

    def test_admin_record_wrappers_support_read_and_delete_behavior(self) -> None:
        self._login()

        self.client.put(
            "/api/admin/employee-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.employee_id,
                "attendance_days": "3",
                "work_hours": "21.5",
                "remark": "员工修正",
            },
        )
        self.client.put(
            "/api/admin/manager-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.manager_id,
                "attendance_days": "20",
                "injury_days": "1",
                "remark": "经理修正",
            },
        )

        employee_get = self.client.get(
            f"/api/admin/employee-attendance-overrides/record?month=2026-05&emp_id={self.employee_id}"
        )
        manager_get = self.client.get(
            f"/api/admin/manager-attendance-overrides/record?month=2026-05&emp_id={self.manager_id}"
        )
        employee_delete = self.client.delete(
            f"/api/admin/employee-attendance-overrides/record?month=2026-05&emp_id={self.employee_id}"
        )
        manager_delete = self.client.delete(
            f"/api/admin/manager-attendance-overrides/record?month=2026-05&emp_id={self.manager_id}"
        )

        self.assertEqual(employee_get.status_code, 200)
        self.assertEqual(employee_get.get_json()["override"]["attendance_days"], 3.0)
        self.assertEqual(manager_get.status_code, 200)
        self.assertEqual(manager_get.get_json()["override"]["injury_days"], 1.0)
        self.assertEqual(employee_delete.status_code, 200)
        self.assertEqual(manager_delete.status_code, 200)
        with self.app.app_context():
            self.assertIsNone(EmployeeAttendanceOverride.query.filter_by(emp_id=self.employee_id).first())
            self.assertIsNone(ManagerAttendanceOverride.query.filter_by(emp_id=self.manager_id).first())
        employee_history = self.client.get(
            "/api/admin/employee-attendance-overrides/history?month=2026-05"
        ).get_json()["rows"]
        manager_history = self.client.get(
            "/api/admin/manager-attendance-overrides/history?month=2026-05"
        ).get_json()["rows"]
        self.assertEqual([row["action_type"] for row in employee_history], ["clear", "manual_save"])
        self.assertEqual([row["action_type"] for row in manager_history], ["clear", "manual_save"])

    def test_admin_attendance_override_import_wrappers_apply_xlsx_files(self) -> None:
        self._login()

        employee_file = self._xlsx_file(
            [
                ["月份", "工号", "姓名", "考勤天数", "工时", "半勤天数", "迟到早退", "备注"],
                ["2026-05", "E001", "员工甲", 3, 21.5, 1, 10, "员工导入"],
            ],
            "employee-overrides.xlsx",
        )
        manager_file = self._xlsx_file(
            [
                ["月份", "工号", "姓名", "出勤天数", "工伤", "出差", "婚假", "丧假", "迟到早退", "备注"],
                ["2026-05", "M001", "经理甲", 20, 1, 2, 0, 0, 5, "经理导入"],
            ],
            "manager-overrides.xlsx",
        )

        employee_response = self.client.post(
            "/api/admin/employee-attendance-overrides/import",
            data={"month": "2026-05", "file": (employee_file, "employee-overrides.xlsx")},
            content_type="multipart/form-data",
        )
        manager_response = self.client.post(
            "/api/admin/manager-attendance-overrides/import",
            data={"month": "2026-05", "file": (manager_file, "manager-overrides.xlsx")},
            content_type="multipart/form-data",
        )

        self.assertEqual(employee_response.status_code, 200)
        self.assertEqual(employee_response.get_json()["changed_count"], 1)
        self.assertEqual(manager_response.status_code, 200)
        self.assertEqual(manager_response.get_json()["changed_count"], 1)
        with self.app.app_context():
            employee_override = EmployeeAttendanceOverride.query.filter_by(emp_id=self.employee_id).first()
            manager_override = ManagerAttendanceOverride.query.filter_by(emp_id=self.manager_id).first()

            self.assertEqual(employee_override.attendance_days, 3.0)
            self.assertEqual(employee_override.work_hours, 21.5)
            self.assertEqual(employee_override.half_days, 1)
            self.assertEqual(employee_override.remark, "员工导入")
            self.assertEqual(manager_override.attendance_days, 20.0)
            self.assertEqual(manager_override.business_trip_days, 2.0)
            self.assertEqual(manager_override.late_early_minutes, 5)
            self.assertEqual(manager_override.remark, "经理导入")

    def test_legacy_admin_dashboard_route_is_not_available(self) -> None:
        self._login()

        response = self.client.get("/admin/dashboard")

        self.assertEqual(response.status_code, 404)


if __name__ == "__main__":
    unittest.main()
